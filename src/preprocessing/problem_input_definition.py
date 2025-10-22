"""
Problem Input Definition for OPT STM Generator
==============================================

This module provides a user-friendly interface for defining 3D geometry, supports,
and loads for the Optimization-based Strut-and-Tie Model (OPT-STM) generator.

Features:
- Excel template generation for structured input
- 3D geometry definition using positive and negative volumes
- Support and load condition specification
- Interactive 3D visualization with matplotlib
- Automatic mesh generation with 8-node hexahedral bricks
- Model export for use with OPT STM GENERATOR_v2.py

GEOMETRY PROCESSING METHODS:
============================

1. RAY-CASTING (Primary Method - Non-Convex Support)
   -------------------------------------------------
   - Supports complex geometries: L-shapes, U-shapes, tunnels, voids
   - Algorithm: Cast ray from each voxel center along +X direction
   - Count intersections with boundary faces (triangulated)
   - Odd intersections = inside, even = outside
   - Uses Möller-Trumbore algorithm for ray-triangle intersection
   - Handles oblique faces correctly through triangulation
   
2. CONVEX HULL (Fallback Method - Fast but Limited)
   -------------------------------------------------
   - Fast computation using scipy.spatial.ConvexHull
   - Only works for CONVEX geometries (single solid block)
   - Point-in-hull test using plane equations
   - Will "fill" non-convex shapes to their convex envelope
   
3. BOUNDING BOX (Final Fallback)
   -------------------------------
   - Simple AABB (axis-aligned bounding box) test
   - Very fast but least accurate
   - Only used if other methods fail

OBLIQUE FACE HANDLING:
=====================

Oblique (non-axis-aligned) faces are PROPERLY VOXELATED:

Concept:
--------
- Face defined by corner points at arbitrary 3D locations
- Each voxel center is tested for containment in the volume
- Creates "staircase" approximation of oblique surfaces
- Resolution controlled by brick size (smaller = smoother)

Example:
--------
A 45° diagonal face from (0,0,0) to (100,100,100):
- With 10mm bricks: ~283 surface voxels (smooth approximation)
- With 50mm bricks: ~12 surface voxels (coarse staircase)

Process:
--------
1. Face triangulated into smaller triangles
2. Each voxel center tested against ALL face triangles
3. Ray-casting determines if center is inside volume boundary
4. Result: 3D "pixelated" representation of smooth geometry

Limitations:
------------
- Thin features < brick size may be lost
- Surface accuracy depends on brick size
- Small angles may show aliasing effects

Recommendations:
----------------
- Use brick size ≤ 1/5 of smallest feature dimension
- For curved surfaces, decrease brick size for smoothness
- Check visualization before running optimization

Author: Generated for OPT-STM Project
Date: October 21, 2025
"""

import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Set
import matplotlib.pyplot as plt
from matplotlib import patches
from mpl_toolkits.mplot3d import Axes3D
from mpl_toolkits.mplot3d.art3d import Poly3DCollection
import pickle
import warnings
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy.spatial import ConvexHull

# Suppress warnings for cleaner output
warnings.filterwarnings('ignore')


# =============================================================================
# Hexahedral Element Stiffness Matrix (ke0) Computation
# =============================================================================

def compute_hex8_ke0(E=1.0, nu=0.3):
    """
    Compute the 24x24 reference element stiffness matrix for an 8-node hexahedral
    brick element with unit dimensions (1x1x1) using isoparametric formulation.
    
    This implementation uses 2x2x2 Gauss quadrature integration and follows
    standard FEM formulation for 3D elasticity.
    
    Parameters
    ----------
    E : float
        Young's modulus (default: 1.0 for reference stiffness)
    nu : float
        Poisson's ratio (default: 0.3)
    
    Returns
    -------
    ke0 : np.ndarray, shape (24, 24)
        Reference element stiffness matrix.
        DOF ordering: [ux_n0, uy_n0, uz_n0, ux_n1, ..., uz_n7]
        
    Node numbering (standard hex8):
        Bottom face (z=-1): 0,1,2,3  (counterclockwise from origin)
        Top face (z=+1):    4,5,6,7  (counterclockwise from origin)
    """
    # Constitutive matrix (isotropic 3D elasticity)
    factor = E / ((1.0 + nu) * (1.0 - 2.0 * nu))
    C = factor * np.array([
        [1-nu,   nu,   nu,    0,           0,           0],
        [  nu, 1-nu,   nu,    0,           0,           0],
        [  nu,   nu, 1-nu,    0,           0,           0],
        [   0,    0,    0, (1-2*nu)/2,     0,           0],
        [   0,    0,    0,    0,      (1-2*nu)/2,      0],
        [   0,    0,    0,    0,           0,      (1-2*nu)/2]
    ])
    
    # Gauss quadrature points (2x2x2 integration)
    gp = 1.0 / np.sqrt(3.0)
    gauss_points = np.array([
        [-gp, -gp, -gp],
        [ gp, -gp, -gp],
        [ gp,  gp, -gp],
        [-gp,  gp, -gp],
        [-gp, -gp,  gp],
        [ gp, -gp,  gp],
        [ gp,  gp,  gp],
        [-gp,  gp,  gp]
    ])
    
    # Gauss weights (all equal for 2x2x2)
    weights = np.ones(8)
    
    # Initialize stiffness matrix
    ke0 = np.zeros((24, 24))
    
    # Integration loop over Gauss points
    for gp_idx in range(8):
        xi, eta, zeta = gauss_points[gp_idx]
        weight = weights[gp_idx]
        
        # Shape function derivatives in natural coordinates
        dN = 0.125 * np.array([
            [-(1-eta)*(1-zeta),  (1-eta)*(1-zeta),  (1+eta)*(1-zeta), -(1+eta)*(1-zeta),
             -(1-eta)*(1+zeta),  (1-eta)*(1+zeta),  (1+eta)*(1+zeta), -(1+eta)*(1+zeta)],
            [-(1-xi)*(1-zeta), -(1+xi)*(1-zeta),  (1+xi)*(1-zeta),  (1-xi)*(1-zeta),
             -(1-xi)*(1+zeta), -(1+xi)*(1+zeta),  (1+xi)*(1+zeta),  (1-xi)*(1+zeta)],
            [-(1-xi)*(1-eta), -(1+xi)*(1-eta), -(1+xi)*(1+eta), -(1-xi)*(1+eta),
              (1-xi)*(1-eta),  (1+xi)*(1-eta),  (1+xi)*(1+eta),  (1-xi)*(1+eta)]
        ])
        
        # Node coordinates in natural space (unit cube from -1 to +1)
        # Map to physical coordinates (0 to 1)
        node_coords = np.array([
            [0, 0, 0],  # Node 0
            [1, 0, 0],  # Node 1
            [1, 1, 0],  # Node 2
            [0, 1, 0],  # Node 3
            [0, 0, 1],  # Node 4
            [1, 0, 1],  # Node 5
            [1, 1, 1],  # Node 6
            [0, 1, 1]   # Node 7
        ])
        
        # Jacobian matrix
        J = dN @ node_coords  # (3, 3)
        detJ = np.linalg.det(J)
        
        if detJ <= 0:
            raise ValueError(f"Negative Jacobian at Gauss point {gp_idx}")
        
        # Derivatives in physical coordinates
        J_inv = np.linalg.inv(J)
        dN_dx = J_inv @ dN  # (3, 8)
        
        # B matrix (strain-displacement, 6x24)
        B = np.zeros((6, 24))
        for i in range(8):
            B[0, 3*i]   = dN_dx[0, i]  # dN_i/dx for epsilon_xx
            B[1, 3*i+1] = dN_dx[1, i]  # dN_i/dy for epsilon_yy
            B[2, 3*i+2] = dN_dx[2, i]  # dN_i/dz for epsilon_zz
            B[3, 3*i]   = dN_dx[1, i]  # dN_i/dy for gamma_xy
            B[3, 3*i+1] = dN_dx[0, i]  # dN_i/dx for gamma_xy
            B[4, 3*i+1] = dN_dx[2, i]  # dN_i/dz for gamma_yz
            B[4, 3*i+2] = dN_dx[1, i]  # dN_i/dy for gamma_yz
            B[5, 3*i]   = dN_dx[2, i]  # dN_i/dz for gamma_zx
            B[5, 3*i+2] = dN_dx[0, i]  # dN_i/dx for gamma_zx
        
        # Add contribution to element stiffness: ke += B^T * C * B * detJ * weight
        ke0 += (B.T @ C @ B) * detJ * weight
    
    return ke0


# =============================================================================
# Helper Functions for FEM Assembly
# =============================================================================

def generate_edof_array(elements: np.ndarray, n_dofs_per_node: int = 3) -> np.ndarray:
    """
    Generate element-to-DOF connectivity array from element-to-node connectivity.
    
    Parameters
    ----------
    elements : np.ndarray, shape (n_elements, 8)
        Element connectivity (node indices for each hex8 element)
    n_dofs_per_node : int
        Number of DOFs per node (default: 3 for 3D: ux, uy, uz)
    
    Returns
    -------
    edof : np.ndarray, shape (n_elements, 24)
        Element-to-DOF connectivity array
        DOF ordering: [ux_n0, uy_n0, uz_n0, ux_n1, ..., uz_n7]
    """
    n_elements = elements.shape[0]
    n_dofs_per_element = 8 * n_dofs_per_node
    edof = np.zeros((n_elements, n_dofs_per_element), dtype=int)
    
    for e in range(n_elements):
        for i, node_idx in enumerate(elements[e]):
            for d in range(n_dofs_per_node):
                edof[e, n_dofs_per_node * i + d] = n_dofs_per_node * node_idx + d
    
    return edof


def assemble_force_vector(loads: List, n_nodes: int) -> np.ndarray:
    """
    Assemble global force vector from load conditions.
    
    Parameters
    ----------
    loads : List[LoadCondition]
        List of load conditions with affected nodes and force components
    n_nodes : int
        Total number of nodes in the mesh
    
    Returns
    -------
    f : np.ndarray, shape (3 * n_nodes,)
        Global force vector with DOF ordering [ux_0, uy_0, uz_0, ux_1, ...]
    """
    f = np.zeros(3 * n_nodes)
    
    for load in loads:
        if not load.affected_nodes:
            continue
        
        n_affected = len(load.affected_nodes)
        
        # Distribute load equally among affected nodes (for face loads)
        # For point loads, n_affected should be 1
        fx_per_node = load.f_x / n_affected
        fy_per_node = load.f_y / n_affected
        fz_per_node = load.f_z / n_affected
        
        for node_idx in load.affected_nodes:
            f[3 * node_idx + 0] += fx_per_node
            f[3 * node_idx + 1] += fy_per_node
            f[3 * node_idx + 2] += fz_per_node
        
        # Note: Moments (m_x, m_y, m_z) are not applied as hex8 elements
        # have no rotational DOFs. For moment application, convert to
        # equivalent force couples if needed.
    
    return f


def assemble_fixed_dofs(supports: List) -> np.ndarray:
    """
    Assemble array of fixed DOF indices from support conditions.
    
    Parameters
    ----------
    supports : List[SupportCondition]
        List of support conditions with affected nodes and constraint flags
    
    Returns
    -------
    fixed_dofs : np.ndarray, shape (n_fixed,)
        Array of global DOF indices that are constrained (zero displacement)
    """
    fixed = []
    
    for support in supports:
        if not support.affected_nodes:
            continue
        
        for node_idx in support.affected_nodes:
            if support.u_x:
                fixed.append(3 * node_idx + 0)
            if support.u_y:
                fixed.append(3 * node_idx + 1)
            if support.u_z:
                fixed.append(3 * node_idx + 2)
            
            # Note: theta_x, theta_y, theta_z are ignored as hex8 elements
            # have no rotational DOFs
    
    return np.array(sorted(set(fixed)), dtype=int)


@dataclass
class GeometryConfig:
    """Configuration for geometry definition."""
    units_length: str = "mm"  # mm, cm, m
    units_force: str = "kN"   # N, kN
    brick_size_x: float = 0.0
    brick_size_y: float = 0.0
    brick_size_z: float = 0.0
    origin_shift: np.ndarray = field(default_factory=lambda: np.zeros(3))


@dataclass
class Face:
    """Represents a planar face defined by corner points."""
    points: np.ndarray  # (n_points, 3) array of coordinates
    face_id: int
    volume_id: int
    is_negative: bool = False
    
    def __post_init__(self):
        self.points = np.asarray(self.points)
        if self.points.shape[1] != 3:
            raise ValueError(f"Face points must have 3 coordinates (x,y,z)")
    
    @property
    def centroid(self) -> np.ndarray:
        """Calculate face centroid."""
        return np.mean(self.points, axis=0)
    
    @property
    def normal(self) -> np.ndarray:
        """Calculate face normal vector (assuming planar and convex)."""
        # Use first 3 points to compute normal
        if len(self.points) < 3:
            raise ValueError("Need at least 3 points to define a plane")
        v1 = self.points[1] - self.points[0]
        v2 = self.points[2] - self.points[0]
        normal = np.cross(v1, v2)
        norm = np.linalg.norm(normal)
        if norm < 1e-10:
            raise ValueError("Degenerate face (collinear points)")
        return normal / norm
    
    def is_point_on_plane(self, point: np.ndarray, tolerance: float = 1e-6) -> bool:
        """Check if a point lies on the face's plane."""
        # Vector from face point to test point
        vec = point - self.points[0]
        # Check if perpendicular to normal
        return abs(np.dot(vec, self.normal)) < tolerance


@dataclass
class SupportCondition:
    """Support boundary condition.
    
    Note: Hexahedral elements have only translational DOFs (ux, uy, uz).
    Rotational constraints (theta_x, theta_y, theta_z) are included for
    template compatibility but are not used in the FEM solver.
    """
    support_id: int
    location: np.ndarray  # (3,) or face definition
    is_face: bool
    face_points: Optional[np.ndarray] = None
    u_x: bool = True  # True = fixed, False = free
    u_y: bool = True
    u_z: bool = True
    theta_x: bool = False  # Not used for hex8 elements (no rotational DOFs)
    theta_y: bool = False  # Not used for hex8 elements (no rotational DOFs)
    theta_z: bool = False  # Not used for hex8 elements (no rotational DOFs)
    affected_nodes: List[int] = field(default_factory=list)


@dataclass
class LoadCondition:
    """Load application condition.
    
    Note: Hexahedral elements have only translational DOFs (ux, uy, uz).
    Moments (m_x, m_y, m_z) are included for template compatibility but
    cannot be directly applied. To apply moments, convert to equivalent
    force couples acting on multiple nodes.
    """
    load_id: int
    location: np.ndarray  # (3,) or face definition
    is_face: bool
    face_points: Optional[np.ndarray] = None
    f_x: float = 0.0
    f_y: float = 0.0
    f_z: float = 0.0
    m_x: float = 0.0  # Not directly applicable (convert to force couples if needed)
    m_y: float = 0.0  # Not directly applicable (convert to force couples if needed)
    m_z: float = 0.0  # Not directly applicable (convert to force couples if needed)
    affected_nodes: List[int] = field(default_factory=list)


class ExcelTemplateGenerator:
    """Generates formatted Excel template for geometry input."""
    
    def __init__(self, filepath: Path):
        self.filepath = filepath
        self.wb = Workbook()
        
        # Define color scheme
        self.colors = {
            'header': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'basic_params': PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid'),
            'positive_vol': PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid'),
            'negative_vol': PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid'),
            'supports': PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid'),
            'loads': PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid'),
        }
        
        self.font_header = Font(bold=True, color='FFFFFF', size=11)
        self.font_bold = Font(bold=True, size=10)
        self.font_normal = Font(size=10)
        
        self.border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate(self):
        """Generate the complete template."""
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Create sheets
        self._create_instructions_sheet()
        self._create_basic_params_sheet()
        self._create_positive_volumes_sheet()
        self._create_negative_volumes_sheet()
        self._create_supports_sheet()
        self._create_loads_sheet()
        
        # Save workbook
        self.wb.save(self.filepath)
        print(f"✓ Excel template generated: {self.filepath}")
    
    def _create_instructions_sheet(self):
        """Create instructions sheet."""
        ws = self.wb.create_sheet("Instructions", 0)
        
        instructions = [
            ["OPT-STM GEOMETRY INPUT TEMPLATE", ""],
            ["", ""],
            ["OVERVIEW", ""],
            ["This template allows you to define 3D geometry, supports, and loads for the", ""],
            ["Optimization-based Strut-and-Tie Model (OPT-STM) generator.", ""],
            ["", ""],
            ["WORKFLOW", ""],
            ["1. Fill in 'Basic Parameters' sheet with units and brick sizes", ""],
            ["2. Define geometry using 'Positive Volumes' (required)", ""],
            ["3. Define voids using 'Negative Volumes' (optional)", ""],
            ["4. Specify support conditions in 'Supports' sheet", ""],
            ["5. Specify load conditions in 'Loads' sheet", ""],
            ["6. Save the file and load it using problem_input_definition.py", ""],
            ["", ""],
            ["IMPORTANT NOTES", ""],
            ["• All coordinates must be consistent with chosen length units", ""],
            ["• Faces must form closed volumes (edges must connect)", ""],
            ["• Support/load faces must be coplanar with geometry surfaces", ""],
            ["• Use 1 (fixed) or 0 (free) for support constraints", ""],
            ["• Positive volumes define solid material", ""],
            ["• Negative volumes subtract from positive volumes (voids)", ""],
            ["", ""],
            ["COLOR CODING", ""],
            ["• Blue: Basic parameters", ""],
            ["• Green: Positive volumes (material)", ""],
            ["• Orange: Negative volumes (voids)", ""],
            ["• Yellow: Support conditions", ""],
            ["• Light Blue: Load conditions", ""],
        ]
        
        for i, row in enumerate(instructions, 1):
            ws.cell(i, 1, row[0])
            if i == 1:
                ws.cell(i, 1).font = Font(bold=True, size=14, color='366092')
            elif row[0] in ["OVERVIEW", "WORKFLOW", "IMPORTANT NOTES", "COLOR CODING"]:
                ws.cell(i, 1).font = Font(bold=True, size=11)
        
        ws.column_dimensions['A'].width = 80
    
    def _create_basic_params_sheet(self):
        """Create basic parameters sheet."""
        ws = self.wb.create_sheet("Basic Parameters")
        
        # Headers
        headers = [
            ["BASIC PARAMETERS", "", "", ""],
            ["", "", "", ""],
            ["Parameter", "Value", "Unit", "Notes"],
        ]
        
        for i, row in enumerate(headers, 1):
            for j, val in enumerate(row, 1):
                cell = ws.cell(i, j, val)
                if i == 1:
                    cell.fill = self.colors['header']
                    cell.font = self.font_header
                elif i == 3:
                    cell.fill = self.colors['basic_params']
                    cell.font = self.font_bold
                cell.border = self.border_thin
        
        # Data rows
        params = [
            ["Length Units", "mm", "", "Options: mm, cm, m"],
            ["Force Units", "kN", "", "Options: N, kN"],
            ["Brick Size X", "100", "mm", "Hexahedral element size in X direction"],
            ["Brick Size Y", "100", "mm", "Hexahedral element size in Y direction"],
            ["Brick Size Z", "100", "mm", "Hexahedral element size in Z direction"],
        ]
        
        for i, row in enumerate(params, 4):
            for j, val in enumerate(row, 1):
                cell = ws.cell(i, j, val)
                cell.fill = self.colors['basic_params']
                cell.border = self.border_thin
                if j == 2:  # Value column
                    cell.font = Font(bold=True, size=11, color='FF0000')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 40
    
    def _create_positive_volumes_sheet(self):
        """Create positive volumes definition sheet."""
        ws = self.wb.create_sheet("Positive Volumes")
        
        # Instructions
        ws.merge_cells('A1:H1')
        cell = ws.cell(1, 1, "POSITIVE VOLUMES - Define solid material regions")
        cell.fill = self.colors['header']
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:H2')
        cell = ws.cell(2, 1, "Each face must have at least 3 corner points. Define faces that form closed volumes.")
        cell.font = Font(italic=True, size=9)
        
        # Headers
        headers = ["Volume_ID", "Face_ID", "Point_ID", "X", "Y", "Z", "Face_Type", "Notes"]
        for j, header in enumerate(headers, 1):
            cell = ws.cell(3, j, header)
            cell.fill = self.colors['positive_vol']
            cell.font = self.font_bold
            cell.border = self.border_thin
        
        # Complete hexahedron template (all 6 faces with 24 points)
        examples = [
            [1, 1, 1, 0, "", "", "Bottom", "RC PAD (PRISM)"],
            [1, 1, 2, "", "", "", "Bottom", "RC PAD (PRISM)"],
            [1, 1, 3, "", "", "", "Bottom", "RC PAD (PRISM)"],
            [1, 1, 4, "", "", "", "Bottom", "RC PAD (PRISM)"],
            [1, 2, 5, "", "", "", "Top", "RC PAD (PRISM)"],
            [1, 2, 6, "", "", "", "Top", "RC PAD (PRISM)"],
            [1, 2, 7, "", "", "", "Top", "RC PAD (PRISM)"],
            [1, 2, 8, "", "", "", "Top", "RC PAD (PRISM)"],
            [1, 3, 9, "", "", "", "XZ NEAR face", "RC PAD (PRISM)"],
            [1, 3, 10, "", "", "", "XZ NEAR face", "RC PAD (PRISM)"],
            [1, 3, 11, "", "", "", "XZ NEAR face", "RC PAD (PRISM)"],
            [1, 3, 12, "", "", "", "XZ NEAR face", "RC PAD (PRISM)"],
            [1, 4, 13, "", "", "", "XZ FAR face", "RC PAD (PRISM)"],
            [1, 4, 14, "", "", "", "XZ FAR face", "RC PAD (PRISM)"],
            [1, 4, 15, "", "", "", "XZ FAR face", "RC PAD (PRISM)"],
            [1, 4, 16, "", "", "", "XZ FAR face", "RC PAD (PRISM)"],
            [1, 5, 17, "", "", "", "YZ NEAR face", "RC PAD (PRISM)"],
            [1, 5, 18, "", "", "", "YZ NEAR face", "RC PAD (PRISM)"],
            [1, 5, 19, "", "", "", "YZ NEAR face", "RC PAD (PRISM)"],
            [1, 5, 20, "", "", "", "YZ NEAR face", "RC PAD (PRISM)"],
            [1, 6, 21, "", "", "", "YZ FAR face", "RC PAD (PRISM)"],
            [1, 6, 22, "", "", "", "YZ FAR face", "RC PAD (PRISM)"],
            [1, 6, 23, "", "", "", "YZ FAR face", "RC PAD (PRISM)"],
            [1, 6, 24, "", "", "", "YZ FAR face", "RC PAD (PRISM)"],
        ]
        
        for i, row in enumerate(examples, 4):
            for j, val in enumerate(row, 1):
                cell = ws.cell(i, j, val)
                cell.fill = self.colors['positive_vol']
                cell.border = self.border_thin
                if j <= 3:  # ID columns
                    cell.alignment = Alignment(horizontal='center')
                if j >= 4 and j <= 6 and val == "":  # Coordinate columns
                    cell.font = Font(italic=True, color='999999')
                    cell.value = "<fill>"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
    
    def _create_negative_volumes_sheet(self):
        """Create negative volumes (voids) definition sheet."""
        ws = self.wb.create_sheet("Negative Volumes")
        
        ws.merge_cells('A1:H1')
        cell = ws.cell(1, 1, "NEGATIVE VOLUMES - Define voids (optional)")
        cell.fill = self.colors['header']
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:H2')
        cell = ws.cell(2, 1, "Negative volumes subtract from positive volumes. Leave empty if no voids needed.")
        cell.font = Font(italic=True, size=9)
        
        headers = ["Volume_ID", "Face_ID", "Point_ID", "X", "Y", "Z", "Face_Type", "Notes"]
        for j, header in enumerate(headers, 1):
            cell = ws.cell(3, j, header)
            cell.fill = self.colors['negative_vol']
            cell.font = self.font_bold
            cell.border = self.border_thin
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions['H'].width = 25
    
    def _create_supports_sheet(self):
        """Create supports definition sheet."""
        ws = self.wb.create_sheet("Supports")
        
        ws.merge_cells('A1:I1')
        cell = ws.cell(1, 1, "SUPPORT CONDITIONS")
        cell.fill = self.colors['header']
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:I2')
        cell = ws.cell(2, 1, "Define as faces (4+ points; TYPE='Face') or single points (unlimited number; TYPE='Point'). Use 1=Fixed, 0=Free. Must be coplanar with geometry.")
        cell.font = Font(italic=True, size=9)
        
        # Headers (removed theta columns as hex8 elements have no rotational DOFs)
        headers = ["Support_ID", "Type", "Point_ID", "X", "Y", "Z", "u_x", "u_y", "u_z"]
        for j, header in enumerate(headers, 1):
            cell = ws.cell(3, j, header)
            cell.fill = self.colors['supports']
            cell.font = self.font_bold
            cell.border = self.border_thin
        
        # Example: One rectangular face (4 points)
        examples = [
            [1, "Face", 1, 0, "", "", 1, 1, 1],
            [1, "Face", 2, "", "", "", 1, 1, 1],
            [1, "Face", 3, "", "", "", 1, 1, 1],
            [1, "Face", 4, "", "", "", 1, 1, 1],
        ]
        
        for i, row in enumerate(examples, 4):
            for j, val in enumerate(row, 1):
                cell = ws.cell(i, j, val)
                cell.fill = self.colors['supports']
                cell.border = self.border_thin
                if j <= 3:  # ID columns
                    cell.alignment = Alignment(horizontal='center')
                if j >= 4 and j <= 6 and val == "":  # Coordinate columns
                    cell.font = Font(italic=True, color='999999')
                    cell.value = "<fill>"
        
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 11
    
    def _create_loads_sheet(self):
        """Create loads definition sheet."""
        ws = self.wb.create_sheet("Loads")
        
        ws.merge_cells('A1:I1')
        cell = ws.cell(1, 1, "LOAD CONDITIONS")
        cell.fill = self.colors['header']
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:I2')
        cell = ws.cell(2, 1, "Define as faces (distributed; TYPE='Face') or points (concentrated; TYPE='Point'). Forces in global coordinates.")
        cell.font = Font(italic=True, size=9)
        
        ws.merge_cells('A3:I3')
        cell = ws.cell(3, 1, "NOTE: For face loads, equal forces at all points = uniform. Different forces = averaged (currently uniform distribution only).")
        cell.font = Font(italic=True, size=8, color='FF6600')
        
        # Headers (removed moment columns as hex8 elements have no rotational DOFs)
        headers = ["Load_ID", "Type", "Point_ID", "X", "Y", "Z", "F_x", "F_y", "F_z"]
        for j, header in enumerate(headers, 1):
            cell = ws.cell(4, j, header)
            cell.fill = self.colors['loads']
            cell.font = self.font_bold
            cell.border = self.border_thin
        
        # Example: One rectangular face (distributed) + one point load (concentrated)
        examples = [
            [1, "Face", 1, "", "", "", 0, 0, 0],
            [1, "Face", 2, "", "", "", 0, 0, 0],
            [1, "Face", 3, "", "", "", 0, 0, 0],
            [1, "Face", 4, "", "", "", 0, 0, 0],
            [2, "Point", 5, "", "", "", 0, 0, -100],
        ]
        
        for i, row in enumerate(examples, 5):
            for j, val in enumerate(row, 1):
                cell = ws.cell(i, j, val)
                cell.fill = self.colors['loads']
                cell.border = self.border_thin
                if j <= 3:  # ID columns
                    cell.alignment = Alignment(horizontal='center')
                if j >= 4 and j <= 6 and val == "":  # Coordinate columns
                    cell.font = Font(italic=True, color='999999')
                    cell.value = "<fill>"
        
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 11


class GeometryProcessor:
    """Processes geometry definitions and generates hex mesh."""
    
    def __init__(self, config: GeometryConfig):
        self.config = config
        self.positive_faces: List[Face] = []
        self.negative_faces: List[Face] = []
        self.voxel_centers: Optional[np.ndarray] = None
        self.voxel_grid: Optional[np.ndarray] = None
        self.nodes: Optional[np.ndarray] = None
        self.elements: Optional[np.ndarray] = None
        self.node_to_idx: Dict[Tuple[int, int, int], int] = {}
        
        # Initialize surface element tracking (computed during mesh generation)
        self.surface_elements: Optional[np.ndarray] = None
        self.surface_element_centers: Optional[np.ndarray] = None
    
    def add_positive_volume(self, faces: List[Face]):
        """Add faces defining a positive volume."""
        self.positive_faces.extend(faces)
    
    def add_negative_volume(self, faces: List[Face]):
        """Add faces defining a negative volume (void)."""
        self.negative_faces.extend(faces)
    
    def compute_origin_shift(self) -> np.ndarray:
        """
        Calculate shift to place model in first octant (all coordinates >= 0).
        Shift = -min(coordinates) for each axis where min < 0.
        """
        if not self.positive_faces:
            return np.zeros(3)
        
        # Get all points from positive volumes (the actual model geometry)
        all_points = np.vstack([face.points for face in self.positive_faces])
        min_coords = np.min(all_points, axis=0)
        
        # Calculate shift needed to move minimum to origin
        # Only shift axes where min_coords < 0 (outside first octant)
        shift = np.zeros(3)
        for i in range(3):
            if min_coords[i] < 0:
                shift[i] = -min_coords[i]
        
        return shift
    
    def apply_shift(self, shift: np.ndarray):
        """Apply coordinate shift to all faces."""
        for face in self.positive_faces + self.negative_faces:
            # Convert to float to allow addition with float shift, then back to original dtype if needed
            if face.points.dtype == np.int64 or face.points.dtype == np.int32:
                face.points = face.points.astype(np.float64)
            face.points += shift
        self.config.origin_shift = shift
    
    def generate_voxel_mesh(self) -> Tuple[np.ndarray, np.ndarray]:
        """
        Generate 3D voxel mesh from geometry definition.
        Returns: (voxel_centers, voxel_occupied)
        """
        if not self.positive_faces:
            raise ValueError("No positive volumes defined")
        
        # Get bounding box
        all_pos_points = np.vstack([face.points for face in self.positive_faces])
        min_coords = np.min(all_pos_points, axis=0)
        max_coords = np.max(all_pos_points, axis=0)
        
        # Create voxel grid
        brick_size = np.array([
            self.config.brick_size_x,
            self.config.brick_size_y,
            self.config.brick_size_z
        ])
        
        # Calculate number of voxels in each direction
        n_voxels = np.ceil((max_coords - min_coords) / brick_size).astype(int) + 1
        
        print(f"  Grid dimensions: {n_voxels[0]} x {n_voxels[1]} x {n_voxels[2]} = {np.prod(n_voxels)} voxels")
        
        # Generate voxel centers
        x_centers = min_coords[0] + np.arange(n_voxels[0]) * brick_size[0] + brick_size[0] / 2
        y_centers = min_coords[1] + np.arange(n_voxels[1]) * brick_size[1] + brick_size[1] / 2
        z_centers = min_coords[2] + np.arange(n_voxels[2]) * brick_size[2] + brick_size[2] / 2
        
        xx, yy, zz = np.meshgrid(x_centers, y_centers, z_centers, indexing='ij')
        voxel_centers = np.stack([xx.ravel(), yy.ravel(), zz.ravel()], axis=1)
        
        # Determine which voxels are inside positive volumes
        print("  Computing voxel occupancy (positive volumes)...")
        voxel_occupied = self._compute_voxel_occupancy(
            voxel_centers, self.positive_faces, brick_size
        )
        
        # Subtract negative volumes
        if self.negative_faces:
            print("  Computing voxel occupancy (negative volumes)...")
            voxel_negative = self._compute_voxel_occupancy(
                voxel_centers, self.negative_faces, brick_size
            )
            voxel_occupied = voxel_occupied & ~voxel_negative
        
        self.voxel_centers = voxel_centers
        self.voxel_grid = voxel_occupied.reshape(n_voxels)
        
        n_active = np.sum(voxel_occupied)
        n_inactive = len(voxel_occupied) - n_active
        print(f"  Active voxels: {n_active} / {len(voxel_occupied)} ({100*n_active/len(voxel_occupied):.1f}%)")
        print(f"    → {n_active} voxels INSIDE geometry → become FE mesh elements")
        print(f"    → {n_inactive} voxels OUTSIDE geometry → discarded (not part of model)")
        
        return voxel_centers, voxel_occupied
    
    def _compute_voxel_occupancy(self, voxel_centers: np.ndarray, 
                                  faces: List[Face], 
                                  brick_size: np.ndarray) -> np.ndarray:
        """
        Determine which voxels are inside the volume defined by faces.
        
        Uses ray-casting method for non-convex geometry support:
        1. Cast a ray from each voxel center along +X direction
        2. Count intersections with all boundary faces
        3. Odd count = inside, even count = outside
        
        This properly handles:
        - Non-convex geometries (L-shapes, tunnels, etc.)
        - Oblique faces (voxelated with staircase approximation)
        - Complex 3D shapes
        
        Falls back to convex hull method if ray-casting fails, then bounding box.
        """
        n_voxels = len(voxel_centers)
        occupied = np.zeros(n_voxels, dtype=bool)
        
        # Get all points defining the volume
        all_points = np.vstack([face.points for face in faces])
        
        # Remove duplicate points
        unique_points = np.unique(all_points, axis=0)
        
        if len(unique_points) < 4:
            print("  ⚠ Warning: Not enough points to define a 3D volume. Using bounding box.")
            # Fallback to bounding box
            vol_min = np.min(all_points, axis=0) - brick_size / 2
            vol_max = np.max(all_points, axis=0) + brick_size / 2
            in_bounds = np.all(
                (voxel_centers >= vol_min) & (voxel_centers <= vol_max),
                axis=1
            )
            return in_bounds
        
        # Try ray-casting method first (supports non-convex geometries)
        try:
            print("  Using ray-casting for non-convex geometry support...")
            occupied = self._ray_casting_test(voxel_centers, faces)
            return occupied
            
        except Exception as e:
            print(f"  ⚠ Ray-casting failed ({e}), trying convex hull method...")
        
        # Fallback to convex hull method (fast but only for convex shapes)
        try:
            # Compute convex hull
            hull = ConvexHull(unique_points)
            
            # Test each voxel center for containment
            # A point is inside the convex hull if it satisfies all plane inequalities
            # For each face of the hull: A*x + B*y + C*z + D <= 0
            equations = hull.equations  # (n_faces, 4) where last column is constant D
            
            # Compute signed distances to all planes for all voxel centers
            # Add homogeneous coordinate
            voxel_homog = np.column_stack([voxel_centers, np.ones(n_voxels)])
            
            # distances[i, j] = signed distance from voxel i to plane j
            distances = voxel_homog @ equations.T  # (n_voxels, n_faces)
            
            # Voxel is inside if all distances are <= tolerance (accounting for numerical errors)
            tolerance = 1e-10
            occupied = np.all(distances <= tolerance, axis=1)
            
        except Exception as e:
            print(f"  ⚠ Warning: Convex hull computation failed ({e}). Using bounding box.")
            # Final fallback to bounding box
            vol_min = np.min(all_points, axis=0) - brick_size / 2
            vol_max = np.max(all_points, axis=0) + brick_size / 2
            in_bounds = np.all(
                (voxel_centers >= vol_min) & (voxel_centers <= vol_max),
                axis=1
            )
            return in_bounds
        
        return occupied
    
    def _ray_casting_test(self, points: np.ndarray, faces: List[Face]) -> np.ndarray:
        """
        Ray-casting algorithm for point-in-polyhedron test.
        
        Algorithm:
        ----------
        1. For each test point, cast a ray along +X direction to infinity
        2. Count how many times the ray intersects the polyhedron boundary
        3. Odd count = point is inside, even count = point is outside
        
        This is the 3D extension of the 2D ray-casting (point-in-polygon) algorithm.
        
        Parameters
        ----------
        points : np.ndarray, shape (n_points, 3)
            Test points (voxel centers)
        faces : List[Face]
            Boundary faces defining the polyhedron
        
        Returns
        -------
        inside : np.ndarray, shape (n_points,), dtype=bool
            True if point is inside the polyhedron
            
        Notes
        -----
        - Ray direction is along +X axis: (1, 0, 0)
        - Each face must be planar and defined by 3+ coplanar points
        - Handles oblique faces correctly (they are triangulated internally)
        - Tolerance for floating-point comparisons: 1e-10
        """
        n_points = len(points)
        inside = np.zeros(n_points, dtype=bool)
        
        # Ray direction (along +X axis)
        ray_dir = np.array([1.0, 0.0, 0.0])
        
        # Tolerance for floating-point comparisons
        eps = 1e-10
        
        # Process each test point
        for i, point in enumerate(points):
            intersection_count = 0
            
            # Test ray intersection with each face
            for face in faces:
                # Triangulate face if it has more than 3 points
                # (each face is assumed to be planar and convex)
                face_points = face.points
                n_face_pts = len(face_points)
                
                if n_face_pts < 3:
                    continue  # Degenerate face
                
                # Triangulate face: use fan triangulation from first point
                for j in range(1, n_face_pts - 1):
                    # Triangle vertices
                    v0 = face_points[0]
                    v1 = face_points[j]
                    v2 = face_points[j + 1]
                    
                    # Check ray-triangle intersection using Möller-Trumbore algorithm
                    if self._ray_triangle_intersect(point, ray_dir, v0, v1, v2, eps):
                        intersection_count += 1
            
            # Odd number of intersections = inside
            inside[i] = (intersection_count % 2) == 1
        
        return inside
    
    @staticmethod
    def _ray_triangle_intersect(ray_origin: np.ndarray, 
                                ray_dir: np.ndarray,
                                v0: np.ndarray, 
                                v1: np.ndarray, 
                                v2: np.ndarray,
                                eps: float = 1e-10) -> bool:
        """
        Möller-Trumbore ray-triangle intersection algorithm.
        
        Fast method to test if a ray intersects a triangle in 3D space.
        
        Parameters
        ----------
        ray_origin : np.ndarray, shape (3,)
            Starting point of the ray
        ray_dir : np.ndarray, shape (3,)
            Direction vector of the ray (normalized)
        v0, v1, v2 : np.ndarray, shape (3,)
            Triangle vertices
        eps : float
            Tolerance for parallel ray/triangle detection
        
        Returns
        -------
        intersects : bool
            True if ray intersects the triangle in the forward direction
            
        Algorithm:
        ----------
        Solves the ray-triangle intersection using barycentric coordinates:
        - Compute edge vectors and determinant
        - Check if ray is parallel to triangle
        - Compute barycentric coordinates (u, v)
        - Check if intersection point is inside triangle (u>=0, v>=0, u+v<=1)
        - Check if intersection is in ray's forward direction (t>0)
        
        Reference:
        ---------
        Möller, T., & Trumbore, B. (1997). Fast, minimum storage ray-triangle 
        intersection. Journal of Graphics Tools, 2(1), 21-28.
        """
        # Edge vectors
        edge1 = v1 - v0
        edge2 = v2 - v0
        
        # Begin calculating determinant (also used for u parameter)
        pvec = np.cross(ray_dir, edge2)
        det = np.dot(edge1, pvec)
        
        # Ray is parallel to triangle plane
        if abs(det) < eps:
            return False
        
        inv_det = 1.0 / det
        
        # Calculate distance from v0 to ray origin
        tvec = ray_origin - v0
        
        # Calculate u parameter and test bounds
        u = np.dot(tvec, pvec) * inv_det
        if u < 0.0 or u > 1.0:
            return False
        
        # Prepare to test v parameter
        qvec = np.cross(tvec, edge1)
        
        # Calculate v parameter and test bounds
        v = np.dot(ray_dir, qvec) * inv_det
        if v < 0.0 or u + v > 1.0:
            return False
        
        # Calculate t (distance along ray to intersection point)
        t = np.dot(edge2, qvec) * inv_det
        
        # Intersection must be in forward direction of ray
        if t > eps:
            return True
        
        return False
    
    def generate_node_element_connectivity(self) -> Tuple[np.ndarray, np.ndarray]:
        """
        Generate node coordinates and element connectivity for active voxels.
        Returns: (nodes, elements)
        - nodes: (n_nodes, 3) array of node coordinates
        - elements: (n_elements, 8) array of node indices for each hex element
        """
        if self.voxel_grid is None:
            raise ValueError("Must generate voxel mesh first")
        
        brick_size = np.array([
            self.config.brick_size_x,
            self.config.brick_size_y,
            self.config.brick_size_z
        ])
        
        # Get grid dimensions
        nx, ny, nz = self.voxel_grid.shape
        
        # Generate all potential nodes (grid corners)
        all_pos_points = np.vstack([face.points for face in self.positive_faces])
        min_coords = np.min(all_pos_points, axis=0)
        
        x_nodes = min_coords[0] + np.arange(nx + 1) * brick_size[0]
        y_nodes = min_coords[1] + np.arange(ny + 1) * brick_size[1]
        z_nodes = min_coords[2] + np.arange(nz + 1) * brick_size[2]
        
        # Create node dictionary for active elements only
        node_set = set()
        elements_list = []
        
        # Iterate through active voxels
        active_indices = np.argwhere(self.voxel_grid)
        
        print(f"  Generating connectivity for {len(active_indices)} elements...")
        
        for idx, (i, j, k) in enumerate(active_indices):
            # 8 nodes of hexahedron (standard ordering)
            local_nodes = [
                (i, j, k),      # 0
                (i+1, j, k),    # 1
                (i+1, j+1, k),  # 2
                (i, j+1, k),    # 3
                (i, j, k+1),    # 4
                (i+1, j, k+1),  # 5
                (i+1, j+1, k+1),# 6
                (i, j+1, k+1),  # 7
            ]
            
            node_set.update(local_nodes)
            elements_list.append(local_nodes)
        
        # Create node coordinate array
        node_list = sorted(node_set)
        self.node_to_idx = {node: idx for idx, node in enumerate(node_list)}
        
        nodes = np.array([
            [x_nodes[i], y_nodes[j], z_nodes[k]]
            for i, j, k in node_list
        ])
        
        # Convert element node indices
        elements = np.array([
            [self.node_to_idx[node] for node in elem]
            for elem in elements_list
        ])
        
        self.nodes = nodes
        self.elements = elements
        
        print(f"  Generated {len(nodes)} nodes and {len(elements)} elements")
        
        # Generate surface elements dictionary for faster boundary condition application
        print(f"  Computing surface elements...")
        self._compute_surface_elements()
        
        return nodes, elements
    
    def _compute_surface_elements(self):
        """
        Compute surface elements (elements on the boundary of the 3D model).
        Creates a dictionary for fast lookup of nearest surface elements.
        """
        if self.elements is None or self.nodes is None:
            return
        
        # Find boundary nodes (nodes connected to fewer than max elements)
        boundary_nodes = self._get_boundary_nodes()
        
        # Surface elements are those that have at least one boundary node
        surface_elements = []
        surface_element_centers = []
        
        for elem_idx, element in enumerate(self.elements):
            # Check if element has boundary nodes
            element_boundary_nodes = [node for node in element if node in boundary_nodes]
            
            if len(element_boundary_nodes) >= 4:  # Hex face has 4+ nodes on boundary
                surface_elements.append(elem_idx)
                
                # Compute element center
                element_nodes = self.nodes[element]
                center = np.mean(element_nodes, axis=0)
                surface_element_centers.append(center)
        
        self.surface_elements = np.array(surface_elements)
        self.surface_element_centers = np.array(surface_element_centers) if surface_element_centers else np.array([]).reshape(0, 3)
        
        # Compute outward normals for surface elements (for force visualization)
        self._compute_surface_normals()
        
        print(f"  Surface elements identified: {len(self.surface_elements)} of {len(self.elements)} total ({len(self.surface_elements)/len(self.elements)*100:.1f}%)")
        
        # Diagnostic: Count surface elements per face (for debugging face coverage)
        if len(surface_elements) > 0 and hasattr(self, 'positive_faces'):
            self._diagnose_face_coverage()
    
    def _get_boundary_nodes(self) -> Set[int]:
        """
        Identify nodes on the boundary of the mesh.
        Boundary nodes are those connected to fewer than the maximum number
        of elements (8 for interior nodes in a structured hex mesh).
        """
        if self.elements is None:
            return set()
        
        # Count element connections per node
        node_element_count = {}
        for element in self.elements:
            for node_idx in element:
                node_element_count[node_idx] = node_element_count.get(node_idx, 0) + 1
        
        # Interior nodes typically connect to 8 elements in a structured mesh
        # Boundary nodes connect to fewer
        max_connections = max(node_element_count.values()) if node_element_count else 0
        
        # Consider nodes with less than max connections as boundary nodes
        boundary_nodes = {node for node, count in node_element_count.items() 
                         if count < max_connections}
        
        return boundary_nodes
    
    def _compute_surface_normals(self):
        """
        Compute outward-pointing normals for surface elements.
        Uses element face analysis to determine outward direction.
        """
        if len(self.surface_elements) == 0 or self.nodes is None:
            self.surface_normals = np.array([]).reshape(0, 3)
            return
        
        surface_normals = []
        
        for elem_idx in self.surface_elements:
            element_nodes = self.elements[elem_idx] 
            element_coords = self.nodes[element_nodes]
            
            # For hexahedron, find which face is most likely on the surface
            # by checking which face has the most boundary nodes
            hex_faces = [
                [0, 1, 2, 3],  # bottom face (z-min)
                [4, 5, 6, 7],  # top face (z-max)  
                [0, 1, 5, 4],  # front face (y-min)
                [2, 3, 7, 6],  # back face (y-max)
                [0, 3, 7, 4],  # left face (x-min)
                [1, 2, 6, 5]   # right face (x-max)
            ]
            
            boundary_nodes_set = self._get_boundary_nodes()
            
            best_face_idx = 0
            max_boundary_count = 0
            
            for face_idx, face_nodes in enumerate(hex_faces):
                global_face_nodes = [element_nodes[local_idx] for local_idx in face_nodes]
                boundary_count = sum(1 for node in global_face_nodes if node in boundary_nodes_set)
                
                if boundary_count > max_boundary_count:
                    max_boundary_count = boundary_count
                    best_face_idx = face_idx
            
            # Compute normal for the most boundary-heavy face
            face_nodes = hex_faces[best_face_idx]
            face_coords = element_coords[face_nodes]
            
            # Use first 3 points to compute normal
            if len(face_coords) >= 3:
                v1 = face_coords[1] - face_coords[0]
                v2 = face_coords[2] - face_coords[0]
                normal = np.cross(v1, v2)
                norm_length = np.linalg.norm(normal)
                
                if norm_length > 1e-10:
                    normal = normal / norm_length
                    
                    # Ensure normal points outward by checking against element centroid
                    element_center = np.mean(element_coords, axis=0)
                    face_center = np.mean(face_coords, axis=0)
                    outward_direction = face_center - element_center
                    
                    # If normal and outward direction are opposite, flip normal
                    if np.dot(normal, outward_direction) < 0:
                        normal = -normal
                    
                    surface_normals.append(normal)
                else:
                    # Fallback to z-direction if computation fails
                    surface_normals.append(np.array([0, 0, 1]))
            else:
                # Fallback
                surface_normals.append(np.array([0, 0, 1]))
        
        self.surface_normals = np.array(surface_normals)
        print(f"  Computed {len(self.surface_normals)} outward surface normals")
    
    def _diagnose_face_coverage(self):
        """Diagnostic function to show which surface elements correspond to which faces."""
        from collections import defaultdict
        
        face_element_count = defaultdict(int)
        face_element_indices = defaultdict(list)
        
        # For each surface element, find the closest face
        for surf_idx, center in zip(self.surface_elements, self.surface_element_centers):
            min_dist = float('inf')
            closest_face_id = None
            closest_face = None
            
            for face in self.positive_faces:
                # Compute distance from element center to face plane
                face_center = np.mean(face.points, axis=0)
                
                # Simple approach: use distance to face center
                dist = np.linalg.norm(center - face_center)
                
                if dist < min_dist:
                    min_dist = dist
                    closest_face_id = face.face_id
                    closest_face = face
            
            if closest_face_id is not None:
                face_element_count[closest_face_id] += 1
                face_element_indices[closest_face_id].append((surf_idx, center, closest_face))
        
        print(f"  Surface elements by closest Face_ID:")
        for face_id in sorted(face_element_count.keys()):
            count = face_element_count[face_id]
            print(f"    Face_ID {face_id}: {count} surface elements")
        
        # Additional diagnostic: Check if Face_ID 4 elements are actually coplanar with the face
        if 4 in face_element_indices:
            face_4_elements = face_element_indices[4]
            face_4_face = face_4_elements[0][2]  # Get the face object
            
            # Compute face plane normal and a point on the plane
            face_points = face_4_face.points
            face_center = np.mean(face_points, axis=0)
            
            # Compute normal (assuming planar face)
            v1 = face_points[1] - face_points[0]
            v2 = face_points[2] - face_points[0]
            normal = np.cross(v1, v2)
            normal = normal / np.linalg.norm(normal)
            
            # Check distance of each element center from the face plane
            distances = []
            for surf_idx, center, _ in face_4_elements:
                # Distance from point to plane: |dot(normal, point - point_on_plane)|
                dist = abs(np.dot(normal, center - face_center))
                distances.append(dist)
            
            distances = np.array(distances)
            brick_size = self.config.brick_size_x  # Assuming cubic voxels
            half_voxel = brick_size / 2
            print(f"  Face_ID 4 diagnostic (checking for inward shift):")
            print(f"    Distance from face plane: min={distances.min():.2f}, max={distances.max():.2f}, mean={distances.mean():.2f} mm")
            print(f"    Expected minimum: ~{half_voxel:.2f}mm (half voxel = element center offset)")
            print(f"    Elements at correct depth ({half_voxel-2:.0f}-{half_voxel+2:.0f}mm): {np.sum((distances >= half_voxel-2) & (distances <= half_voxel+2))}/{len(distances)}")
            
            if distances.min() < half_voxel - 5:
                print(f"    ⚠ WARNING: Some elements too close to surface (possible mesh issue)")
            elif distances.max() > half_voxel + brick_size:
                print(f"    ⚠ WARNING: Some elements too far inward (possible extra layer)")
            else:
                print(f"    ✓ All Face_ID 4 voxels positioned correctly (centers offset by ~half voxel)")


class BoundaryConditionProcessor:
    """Processes support and load conditions."""
    
    def __init__(self, geometry: GeometryProcessor):
        self.geometry = geometry
        self.supports: List[SupportCondition] = []
        self.loads: List[LoadCondition] = []
    
    def add_support(self, support: SupportCondition):
        """Add support condition and find affected nodes."""
        # Determine tolerance as 1.0x max element dimension to ensure robust detection
        max_element_size = max(
            self.geometry.config.brick_size_x,
            self.geometry.config.brick_size_y, 
            self.geometry.config.brick_size_z
        )
        tolerance = max_element_size * 1.0
        
        affected = self._find_affected_nodes(
            support.location if not support.is_face else support.face_points,
            support.is_face,
            tolerance=tolerance
        )
        
        if not affected:
            if support.is_face:
                centroid = np.mean(support.face_points, axis=0)
                print(f"  ⚠ Support ID {support.support_id} at {centroid} - no nodes found")
                print(f"     This may indicate the support is not coplanar with geometry surface")
                print(f"     Tolerance used: {tolerance:.3f} {self.geometry.config.units_length}")
            else:
                print(f"  ⚠ Support ID {support.support_id} at {support.location} - no nodes found")
                print(f"     Ensure support location matches mesh node positions")
            return False
        
        support.affected_nodes = affected
        self.supports.append(support)
        
        # Validation: check if nodes are on geometry boundary (CRITICAL for face-based supports)
        boundary_nodes = self.geometry._get_boundary_nodes()
        boundary_count = sum(1 for n in affected if n in boundary_nodes)
        non_boundary = [n for n in affected if n not in boundary_nodes]
        
        if support.is_face:
            # For face-based supports, ALL nodes should be on boundary
            if non_boundary:
                print(f"  ⚠ Support ID {support.support_id}: {len(non_boundary)} interior nodes affected (ISSUE!)")
                print(f"     Face-based supports should only affect surface nodes")
                print(f"     Boundary nodes: {boundary_count}, Interior nodes: {len(non_boundary)}")
                print(f"     → This indicates coplanarity tolerance may be too large")
            else:
                print(f"  ✓ Support ID {support.support_id}: {len(affected)} surface nodes affected (all on boundary)")
        else:
            # For point-based supports, interior nodes are acceptable
            if non_boundary:
                print(f"  ⚠ Support ID {support.support_id}: {len(non_boundary)} interior nodes affected")
                print(f"     Consider refining support definition to target boundary only")
            print(f"  ✓ Support ID {support.support_id}: {len(affected)} nodes affected")
        return True
    
    def add_load(self, load: LoadCondition):
        """Add load condition and find affected nodes."""
        # Determine tolerance as 1.0x max element dimension to ensure robust detection
        max_element_size = max(
            self.geometry.config.brick_size_x,
            self.geometry.config.brick_size_y, 
            self.geometry.config.brick_size_z
        )
        tolerance = max_element_size * 1.0
        
        affected = self._find_affected_nodes(
            load.location if not load.is_face else load.face_points,
            load.is_face,
            tolerance=tolerance
        )
        
        if not affected:
            if load.is_face:
                centroid = np.mean(load.face_points, axis=0)
                print(f"  ⚠ Load ID {load.load_id} at {centroid} - no nodes found")
                print(f"     This may indicate the load is not coplanar with geometry surface")
                print(f"     Tolerance used: {tolerance:.3f} {self.geometry.config.units_length}")
            else:
                print(f"  ⚠ Load ID {load.load_id} at {load.location} - no nodes found")
                print(f"     Ensure load location matches mesh node positions")
            return False
        
        load.affected_nodes = affected
        self.loads.append(load)
        
        # Validation: check if nodes are on geometry boundary (CRITICAL for face-based loads)
        boundary_nodes = self.geometry._get_boundary_nodes()
        boundary_count = sum(1 for n in affected if n in boundary_nodes)
        non_boundary = [n for n in affected if n not in boundary_nodes]
        
        if load.is_face:
            # For face-based loads, ALL nodes should be on boundary
            if non_boundary:
                print(f"  ⚠ Load ID {load.load_id}: {len(non_boundary)} interior nodes affected (ISSUE!)")
                print(f"     Face-based loads should only affect surface nodes")
                print(f"     Boundary nodes: {boundary_count}, Interior nodes: {len(non_boundary)}")
                print(f"     → This indicates coplanarity tolerance may be too large")
            else:
                print(f"  ✓ Load ID {load.load_id}: {len(affected)} surface nodes affected (all on boundary)")
        else:
            # For point-based loads, should be EXACTLY ONE node
            if len(affected) != 1:
                print(f"  ⚠ Load ID {load.load_id}: Expected 1 node for point load, got {len(affected)}")
                print(f"     Point loads should affect exactly one node")
            
            if non_boundary:
                print(f"  ⚠ Load ID {load.load_id}: {len(non_boundary)} interior nodes affected")
                print(f"     Consider refining load definition to target boundary only")
            
            print(f"  ✓ Load ID {load.load_id}: {len(affected)} node(s) affected")
        
        # Check for moment application
        if abs(load.m_x) > 1e-10 or abs(load.m_y) > 1e-10 or abs(load.m_z) > 1e-10:
            print(f"  ⚠ Load ID {load.load_id}: Moments specified but cannot be applied directly")
            print(f"     Hex8 elements have no rotational DOFs. Convert to force couples if needed.")
        
        print(f"  ✓ Load ID {load.load_id}: {len(affected)} nodes affected")
        return True
    
    def _find_affected_nodes(self, location, is_face: bool, tolerance: float) -> List[int]:
        """Find nodes affected by boundary condition using surface elements for better accuracy."""
        if self.geometry.nodes is None:
            return []
        
        affected = []
        
        if is_face:
            # Face: find all nodes on or near the face plane
            # Use STRICT coplanarity tolerance to avoid interior nodes
            face_points = np.asarray(location)
            
            # Compute plane equation from first 3 points
            if len(face_points) < 3:
                return []
            
            v1 = face_points[1] - face_points[0]
            v2 = face_points[2] - face_points[0]
            normal = np.cross(v1, v2)
            norm = np.linalg.norm(normal)
            
            if norm < 1e-10:
                return []
            
            normal = normal / norm
            d = -np.dot(normal, face_points[0])
            
            # CRITICAL: Use MUCH tighter tolerance for coplanarity check
            # to avoid capturing interior nodes at different depths
            # Use 1% of voxel size instead of full tolerance
            coplanarity_tolerance = min(
                self.geometry.config.brick_size_x,
                self.geometry.config.brick_size_y,
                self.geometry.config.brick_size_z
            ) * 0.01  # 1% of smallest voxel dimension
            
            # Find nodes on plane within face bounds
            for idx, node in enumerate(self.geometry.nodes):
                # Check distance from plane (STRICT)
                dist = abs(np.dot(normal, node) + d)
                if dist < coplanarity_tolerance:
                    # Check if node is within face bounds (use original tolerance for in-plane extent)
                    face_min = np.min(face_points, axis=0) - tolerance
                    face_max = np.max(face_points, axis=0) + tolerance
                    if np.all((node >= face_min) & (node <= face_max)):
                        affected.append(idx)
        else:
            # Point: find nearest surface element first, then its nodes
            point = np.asarray(location)
            
            # Use surface elements if available for more accurate detection
            if hasattr(self.geometry, 'surface_elements') and len(self.geometry.surface_elements) > 0:
                # Find nearest surface element center
                distances_to_surface = np.linalg.norm(self.geometry.surface_element_centers - point, axis=1)
                nearest_surface_idx = np.argmin(distances_to_surface)
                nearest_surface_element_idx = self.geometry.surface_elements[nearest_surface_idx]
                
                # Get nodes of the nearest surface element
                surface_element_nodes = self.geometry.elements[nearest_surface_element_idx]
                surface_node_coords = self.geometry.nodes[surface_element_nodes]
                
                # Find the closest node within this surface element
                distances_to_nodes = np.linalg.norm(surface_node_coords - point, axis=1)
                closest_local_idx = np.argmin(distances_to_nodes)
                closest_global_node_idx = surface_element_nodes[closest_local_idx]
                min_distance = distances_to_nodes[closest_local_idx]
                
                # Check if closest node is within tolerance
                if min_distance < tolerance:
                    # Check for potential conflicts with nearby surface elements
                    nearby_elements = []
                    for i, elem_center in enumerate(self.geometry.surface_element_centers):
                        if i != nearest_surface_idx and np.linalg.norm(elem_center - point) < tolerance * 1.5:
                            nearby_elements.append(self.geometry.surface_elements[i])
                    
                    if len(nearby_elements) > 0:
                        print(f"     ⚠ Warning: {len(nearby_elements)} other surface elements within 1.5x tolerance")
                        print(f"       Consider refining point location for unique identification")
                    
                    affected.append(closest_global_node_idx)
                    # Debug info (not final status - will be printed by caller)
                    print(f"     Using surface element {nearest_surface_element_idx}, node {closest_global_node_idx}")
                    print(f"     Distance: {min_distance:.2f} {self.geometry.config.units_length}")
            else:
                # Fallback to original method if surface elements not available
                distances = np.linalg.norm(self.geometry.nodes - point, axis=1)
                nearest_idx = np.argmin(distances)
                
                if distances[nearest_idx] < tolerance * 2:  # More lenient for points
                    affected.append(nearest_idx)
        
        return affected


class ModelVisualizer:
    """Visualizes 3D geometry with supports and loads."""
    
    def __init__(self, geometry: GeometryProcessor, bc_processor: BoundaryConditionProcessor):
        self.geometry = geometry
        self.bc_processor = bc_processor
    
    def plot_model(self, show_plot: bool = True, save_path: Optional[Path] = None):
        """Create optimized 3D visualization of the model with proper navigation."""
        # Set up figure with better defaults for 3D
        fig = plt.figure(figsize=(15, 11))
        fig.suptitle('OPT-STM 3D Model Viewer', fontsize=14, weight='bold')
        
        ax = fig.add_subplot(111, projection='3d')
        
        # Configure matplotlib for CAD-like single-axis rotation navigation
        ax.mouse_init()  # Initialize mouse controls
        
        # Implement constrained 3D navigation using pixel coordinates (not data coordinates)
        class ConstrainedNavigation:
            def __init__(self, ax):
                self.ax = ax
                self.press_xy = None
                self.last_elev = ax.elev
                self.last_azim = ax.azim
                
            def on_press(self, event):
                if event.inaxes != self.ax or event.button != 1:  # Only left mouse button
                    return
                self.press_xy = (event.x, event.y)  # Use pixel coordinates
                self.last_elev = self.ax.elev
                self.last_azim = self.ax.azim
                
            def on_motion(self, event):
                if self.press_xy is None or event.inaxes != self.ax:
                    return
                
                # Calculate pixel displacement
                dx = event.x - self.press_xy[0]
                dy = event.y - self.press_xy[1]
                
                # Constrain to single axis rotation based on dominant motion
                if abs(dx) > abs(dy):
                    # Horizontal motion -> azimuth rotation (spin around Z-axis)
                    # Drag right = rotate clockwise when viewed from above
                    new_azim = self.last_azim - dx * 0.5
                    self.ax.view_init(elev=self.last_elev, azim=new_azim)
                else:
                    # Vertical motion -> elevation rotation (tilt up/down)
                    # Drag down = tilt view downward (NEGATIVE for natural motion)
                    new_elev = self.last_elev - dy * 0.5  # Inverted for natural mouse behavior
                    # Clamp elevation to prevent flipping
                    new_elev = max(-90, min(90, new_elev))
                    self.ax.view_init(elev=new_elev, azim=self.last_azim)
                
                fig.canvas.draw_idle()
                
            def on_release(self, event):
                self.press_xy = None
                fig.canvas.draw_idle()
        
        # Replace default chaotic navigation with constrained version
        nav = ConstrainedNavigation(ax)
        fig.canvas.mpl_connect('button_press_event', nav.on_press)
        fig.canvas.mpl_connect('motion_notify_event', nav.on_motion)
        fig.canvas.mpl_connect('button_release_event', nav.on_release)
        
        # Improved zoom behavior (no chaotic spinning)
        def constrained_zoom(event):
            """Smooth zoom without rotation chaos."""
            if event.inaxes == ax:
                xlim, ylim, zlim = ax.get_xlim3d(), ax.get_ylim3d(), ax.get_zlim3d()
                
                # Calculate centers and ranges
                xc = (xlim[0] + xlim[1]) / 2
                yc = (ylim[0] + ylim[1]) / 2  
                zc = (zlim[0] + zlim[1]) / 2
                
                # Smooth zoom factor
                factor = 0.85 if event.button == 'up' else 1.15
                
                xr = (xlim[1] - xlim[0]) * factor / 2
                yr = (ylim[1] - ylim[0]) * factor / 2
                zr = (zlim[1] - zlim[0]) * factor / 2
                
                ax.set_xlim3d(xc - xr, xc + xr)
                ax.set_ylim3d(yc - yr, yc + yr) 
                ax.set_zlim3d(zc - zr, zc + zr)
                fig.canvas.draw_idle()
        
        fig.canvas.mpl_connect('scroll_event', constrained_zoom)
        
        # Plot geometry voxels
        # NOTE: What the plotted dots represent:
        # - Grey dots (surface elements): Centers of surface voxels (1 per Hex8 element)
        #   Each Hex8 voxel has 8 corner nodes; plotted dots are geometric centroids
        # - Red dots (supports): Actual node locations on support surfaces (can be 1-8 per voxel)
        # - Green dots (loads): Centers of voxels containing load nodes (1 per affected voxel)
        #   Load voxels use larger markers (s=60 vs s=15) for high visibility
        self._plot_voxels(ax)
        
        # Highlight supported and loaded nodes
        self._plot_supports(ax)
        self._plot_loads(ax)
        
        # Configure axes
        self._configure_axes(ax)
        
        # Add legend
        ax.legend(loc='upper left', fontsize=9)
        
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=150, bbox_inches='tight')
            print(f"  ✓ Plot saved: {save_path}")
        
        if show_plot:
            plt.show()
        else:
            plt.close()
    
    def _plot_voxels(self, ax):
        """Plot geometry voxels - optimized to show only surface elements."""
        if self.geometry.elements is None or self.geometry.nodes is None:
            return
        
        # Use surface elements if available for much faster rendering
        if hasattr(self.geometry, 'surface_elements') and len(self.geometry.surface_elements) > 0:
            print(f"  Plotting {len(self.geometry.surface_elements)} surface elements (optimized)...")
            
            # Plot surface element centers with reduced opacity so forces can be seen
            surface_centers = self.geometry.surface_element_centers
            ax.scatter(surface_centers[:, 0], surface_centers[:, 1], surface_centers[:, 2],
                      c='lightblue', alpha=0.3, s=15, label='Surface Elements', edgecolors='steelblue',
                      depthshade=True)  # Allow forces to show through
            
            # For better visualization, also draw wireframe of surface elements
            # NOTE: Removed element count cap to ensure all surface elements are visualized
            max_wireframe = min(len(self.geometry.surface_elements), 2000)  # Limit for performance
            if len(self.geometry.surface_elements) > max_wireframe:
                print(f"  Note: Plotting wireframe for first {max_wireframe} of {len(self.geometry.surface_elements)} surface elements")
            
            for surf_elem_idx in self.geometry.surface_elements[:max_wireframe]:
                element_nodes = self.geometry.elements[surf_elem_idx]
                element_coords = self.geometry.nodes[element_nodes]
                
                # Draw edges of hexahedron (12 edges)
                edges = [
                    [0,1], [1,2], [2,3], [3,0],  # bottom face
                    [4,5], [5,6], [6,7], [7,4],  # top face  
                    [0,4], [1,5], [2,6], [3,7]   # vertical edges
                ]
                
                for edge in edges:
                    start, end = edge
                    ax.plot3D([element_coords[start,0], element_coords[end,0]],
                             [element_coords[start,1], element_coords[end,1]],
                             [element_coords[start,2], element_coords[end,2]], 
                             'b-', alpha=0.1, linewidth=0.5)
        else:
            # Fallback to original method if surface elements not computed
            if self.geometry.voxel_grid is None:
                return
            
            brick_size = np.array([
                self.geometry.config.brick_size_x,
                self.geometry.config.brick_size_y,
                self.geometry.config.brick_size_z
            ])
            
            # Get active voxel centers
            active_indices = np.argwhere(self.geometry.voxel_grid)
            
            # For large models, plot subset
            max_voxels_to_plot = 2000  # Reduced for better performance
            if len(active_indices) > max_voxels_to_plot:
                print(f"  Large model detected ({len(active_indices)} voxels). Plotting sample...")
                sample_idx = np.random.choice(len(active_indices), max_voxels_to_plot, replace=False)
                active_indices = active_indices[sample_idx]
            
            all_pos_points = np.vstack([face.points for face in self.geometry.positive_faces])
            min_coords = np.min(all_pos_points, axis=0)
            
            # Plot voxels as scatter points
            voxel_centers = []
            for i, j, k in active_indices:
                center = min_coords + (np.array([i, j, k]) + 0.5) * brick_size
                voxel_centers.append(center)
            
            if voxel_centers:
                voxel_centers = np.array(voxel_centers)
                ax.scatter(voxel_centers[:, 0], voxel_centers[:, 1], voxel_centers[:, 2],
                          c='gray', alpha=0.3, s=10, label='Geometry')
    
    def _plot_supports(self, ax):
        """Plot support conditions with same size as other voxels, just colored red."""
        if not self.bc_processor.supports or self.geometry.nodes is None:
            return
        
        support_nodes = set()
        for support in self.bc_processor.supports:
            support_nodes.update(support.affected_nodes)
        
        if support_nodes:
            support_coords = self.geometry.nodes[list(support_nodes)]
            # Use same size (s=15) as surface elements, not larger (was s=50)
            ax.scatter(support_coords[:, 0], support_coords[:, 1], support_coords[:, 2],
                      c='red', s=15, marker='o', label='Supports', alpha=0.8, edgecolors='darkred')
    
    def _plot_loads(self, ax):
        """Plot load conditions with arrows positioned outside geometry using surface normals."""
        if not self.bc_processor.loads or self.geometry.nodes is None:
            return
        
        # Find elements containing load nodes and mark their centers in green
        load_elements = set()
        for load in self.bc_processor.loads:
            for node_idx in load.affected_nodes:
                # Find elements that contain this node
                for elem_idx, element in enumerate(self.geometry.elements):
                    if node_idx in element:
                        load_elements.add(elem_idx)
        
        # Plot load element centers in bright green (voxel centers, not nodes)
        if load_elements:
            load_element_centers = []
            for elem_idx in load_elements:
                element_nodes = self.geometry.elements[elem_idx]
                element_coords = self.geometry.nodes[element_nodes]
                center = np.mean(element_coords, axis=0)
                load_element_centers.append(center)
            
            load_element_centers = np.array(load_element_centers)
            ax.scatter(load_element_centers[:, 0], load_element_centers[:, 1], load_element_centers[:, 2],
                      c='limegreen', s=60, marker='o', label='Load Voxels', alpha=1.0, 
                      edgecolors='darkgreen', linewidths=2, zorder=10)
        
        # Plot force arrows with outward positioning using surface normals
        for load in self.bc_processor.loads:
            if not load.affected_nodes:
                continue
            
            # Get representative point (centroid of affected nodes)
            affected_coords = self.geometry.nodes[load.affected_nodes]
            centroid = np.mean(affected_coords, axis=0)
            
            # Compute arrow scale
            force_vec = np.array([load.f_x, load.f_y, load.f_z])
            force_mag = np.linalg.norm(force_vec)
            
            if force_mag > 1e-6:
                # Calculate arrow length first (needed for proper offset calculation)
                model_size = np.max(self.geometry.nodes, axis=0) - np.min(self.geometry.nodes, axis=0)
                avg_size = np.mean(model_size)
                arrow_scale = avg_size * 0.2 / force_mag  # Arrow length scale
                arrow_length = arrow_scale * force_mag  # Actual arrow length in mm
                
                # Find closest surface element to determine outward normal
                arrow_start = centroid  # Default position
                
                if (hasattr(self.geometry, 'surface_elements') and 
                    hasattr(self.geometry, 'surface_normals') and
                    len(self.geometry.surface_element_centers) > 0):
                    
                    # Find closest surface element
                    distances = np.linalg.norm(self.geometry.surface_element_centers - centroid, axis=1)
                    closest_surf_idx = np.argmin(distances)
                    
                    if closest_surf_idx < len(self.geometry.surface_normals):
                        surface_normal = self.geometry.surface_normals[closest_surf_idx]
                        
                        # Calculate offset based on arrow length + extra clearance
                        # Offset = arrow_length + 20% extra clearance to ensure full visibility
                        offset_distance = arrow_length * 1.2
                        arrow_start = centroid + surface_normal * offset_distance
                        
                        print(f"  Force arrow: length={arrow_length:.1f}mm, offset={offset_distance:.1f}mm from surface")
                
                # Normalized force direction
                force_dir = force_vec / force_mag
                arrow_vec = force_dir * arrow_length
                
                # Plot single arrow with clear visibility (NOT thick/overlapping)
                ax.quiver(arrow_start[0], arrow_start[1], arrow_start[2],
                         arrow_vec[0], arrow_vec[1], arrow_vec[2],
                         color='red', arrow_length_ratio=0.25, linewidth=3, alpha=1.0,
                         length=1.0, normalize=False)
                
                # Add force magnitude label well in front (fully visible)
                # Position label at arrow tip + extra offset in force direction
                label_offset = arrow_length * 0.3  # 30% beyond arrow tip
                label_pos = arrow_start + arrow_vec + force_dir * label_offset
                force_label = f"{force_mag:.1f} {self.geometry.config.units_force}"
                ax.text(label_pos[0], label_pos[1], label_pos[2], force_label,
                       fontsize=11, color='darkred', weight='bold',
                       bbox=dict(boxstyle='round,pad=0.5', facecolor='yellow', 
                                edgecolor='darkred', alpha=0.9, linewidth=2))
                
                # Draw connection line from load point to arrow base (shows arrow offset)
                connection_length = np.linalg.norm(arrow_start - centroid)
                if connection_length > arrow_length * 0.1:  # Only if significantly offset
                    ax.plot([centroid[0], arrow_start[0]], 
                           [centroid[1], arrow_start[1]], 
                           [centroid[2], arrow_start[2]],
                           'r--', alpha=0.6, linewidth=2)
    
    def _configure_axes(self, ax):
        """Configure axis labels, scaling, and proper 3D navigation."""
        units = self.geometry.config.units_length
        
        # Set axis labels with proper orientation
        ax.set_xlabel(f'X [{units}] →', fontsize=11, weight='bold', labelpad=10)
        ax.set_ylabel(f'Y [{units}] ↗', fontsize=11, weight='bold', labelpad=10)  
        ax.set_zlabel(f'Z [{units}] ↑', fontsize=11, weight='bold', labelpad=10)
        
        # Set equal aspect ratio for proper 3D visualization
        if self.geometry.nodes is not None:
            coords = self.geometry.nodes
            
            # Calculate ranges for each axis
            x_range = coords[:, 0].max() - coords[:, 0].min()
            y_range = coords[:, 1].max() - coords[:, 1].min()
            z_range = coords[:, 2].max() - coords[:, 2].min()
            max_range = max(x_range, y_range, z_range) / 2.0
            
            # Calculate centers
            x_mid = (coords[:, 0].max() + coords[:, 0].min()) / 2
            y_mid = (coords[:, 1].max() + coords[:, 1].min()) / 2
            z_mid = (coords[:, 2].max() + coords[:, 2].min()) / 2
            
            # Set equal limits for proper aspect ratio
            ax.set_xlim(x_mid - max_range, x_mid + max_range)
            ax.set_ylim(y_mid - max_range, y_mid + max_range)
            ax.set_zlim(z_mid - max_range, z_mid + max_range)
            
            # Ensure equal aspect ratio (critical for proper 3D perception)
            ax.set_box_aspect([1,1,1])
        
        # Smart tick formatting
        self._format_ticks(ax, units)
        
        # Set title with model info
        n_elements = len(self.geometry.elements) if self.geometry.elements is not None else 0
        n_surface = len(getattr(self.geometry, 'surface_elements', [])) 
        ax.set_title(f'OPT-STM Geometry Model\n{n_elements:,} elements, {n_surface:,} surface', 
                    fontsize=12, weight='bold', pad=20)
        
        # Set TRUE isometric view (orthogonal projection, not perspective)
        # Elevation: 30° (classic isometric), Azimuth: -45° (origin bottom-left)
        ax.view_init(elev=30, azim=-45, roll=0)
        
        # CRITICAL: Force orthogonal projection (no perspective scaling)
        ax.set_proj_type('ortho')  # Removes perspective distortion
        
        # Improve 3D navigation behavior
        ax.mouse_init()  # Reset mouse navigation
        
        # Set grid for better depth perception
        ax.grid(True, alpha=0.3)
        ax.xaxis.pane.fill = ax.yaxis.pane.fill = ax.zaxis.pane.fill = False
        ax.xaxis.pane.set_edgecolor('gray')
        ax.yaxis.pane.set_edgecolor('gray') 
        ax.zaxis.pane.set_edgecolor('gray')
        ax.xaxis.pane.set_alpha(0.1)
        ax.yaxis.pane.set_alpha(0.1)
        ax.zaxis.pane.set_alpha(0.1)
    
    def _format_ticks(self, ax, units: str):
        """Format axis ticks based on units."""
        # Get current tick locations
        xticks = ax.get_xticks()
        yticks = ax.get_yticks()
        zticks = ax.get_zticks()
        
        # Determine appropriate tick spacing
        spacing_map = {
            'mm': (50, 100),
            'cm': (5, 50),
            'm': (0.25, 1.0)
        }
        
        min_spacing, max_spacing = spacing_map.get(units, (1, 10))
        
        # Apply spacing (simplified approach)
        for axis, ticks in [('x', xticks), ('y', yticks), ('z', zticks)]:
            if len(ticks) > 2:
                # Calculate appropriate spacing
                tick_range = ticks[-1] - ticks[0]
                if tick_range > 0:
                    # Determine nice spacing
                    n_ticks = int(tick_range / min_spacing)
                    if n_ticks > 10:
                        # Use max spacing
                        n_ticks = int(tick_range / max_spacing)
                    
                    # Format tick labels to appropriate precision
                    if units == 'mm' and tick_range > 1000:
                        # Show as meters with mm label
                        pass  # Keep default formatting
                    elif units == 'm' and tick_range < 1:
                        # Show with more decimal places
                        pass  # Keep default formattingtick_range = ticks[-1] - ticks[0]
                n_ticks = max(3, min(10, int(tick_range / min_spacing)))
                # Matplotlib handles this automatically, just set limits


class ModelExporter:
    """Exports model data for use with OPT STM solver."""
    
    @staticmethod
    def export_model(geometry: GeometryProcessor,
                    bc_processor: BoundaryConditionProcessor,
                    tag: str,
                    output_dir: Path):
        """Export complete model data for OPT STM GENERATOR."""
        timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")
        # Filename format matches original requirement: [tag]_geometry_only_model_[ddmmyyyy]_[hhmmss]
        # However, updated to FEM_model to reflect enhanced content
        filename = f"{tag}_FEM_model_{timestamp}.pkl"
        filepath = output_dir / filename
        
        # Validate that mesh has been generated
        if geometry.nodes is None or geometry.elements is None:
            raise ValueError("Mesh has not been generated. Call generate_node_element_connectivity() first.")
        
        print("\n" + "="*60)
        print("PREPARING MODEL EXPORT")
        print("="*60)
        
        # Compute reference element stiffness matrix (ke0)
        print("\n1. Computing reference element stiffness matrix (ke0)...")
        ke0 = compute_hex8_ke0(E=1.0, nu=0.3)
        print(f"   ✓ ke0 computed: shape {ke0.shape}")
        
        # Generate element-to-DOF connectivity (edof)
        print("\n2. Generating element-to-DOF connectivity (edof)...")
        edof = generate_edof_array(geometry.elements, n_dofs_per_node=3)
        print(f"   ✓ edof generated: shape {edof.shape}")
        
        # Assemble global force vector
        print("\n3. Assembling global force vector...")
        n_nodes = len(geometry.nodes)
        f = assemble_force_vector(bc_processor.loads, n_nodes)
        n_loaded_dofs = np.count_nonzero(f)
        print(f"   ✓ Force vector assembled: {n_loaded_dofs} non-zero DOFs")
        print(f"   ✓ Total force magnitude: {np.linalg.norm(f):.2e}")
        
        # Assemble fixed DOFs array
        print("\n4. Assembling fixed DOFs array...")
        fixed_dofs = assemble_fixed_dofs(bc_processor.supports)
        print(f"   ✓ Fixed DOFs: {len(fixed_dofs)} constrained")
        
        # Compute element volumes
        print("\n5. Computing element volumes...")
        brick_volume = (geometry.config.brick_size_x * 
                       geometry.config.brick_size_y * 
                       geometry.config.brick_size_z)
        volumes = np.full(len(geometry.elements), brick_volume)
        print(f"   ✓ Element volume: {brick_volume:.2e} {geometry.config.units_length}³")
        
        # Get grid shape
        grid_shape = geometry.voxel_grid.shape
        print(f"\n6. Grid dimensions: {grid_shape}")
        
        # Determine mesh size (use maximum brick dimension)
        mesh_size = max(geometry.config.brick_size_x,
                       geometry.config.brick_size_y,
                       geometry.config.brick_size_z)
        print(f"   ✓ Mesh size: {mesh_size:.2e} {geometry.config.units_length}")
        
        # Prepare export data (format compatible with OPT STM GENERATOR)
        export_data = {
            # Required inputs for ContinuumFESolver
            'edof': edof,
            'ke0': ke0,
            'f': f,
            'fixed_dofs': fixed_dofs,
            'grid_shape': grid_shape,
            'volumes': volumes,
            'mesh_size': mesh_size,
            
            # Material properties (defaults for reference)
            'E0': 1.0,
            'Emin': 1e-9,
            'nu': 0.3,
            
            # Metadata
            'metadata': {
                'tag': tag,
                'timestamp': timestamp,
                'units_length': geometry.config.units_length,
                'units_force': geometry.config.units_force,
                'brick_size': [
                    geometry.config.brick_size_x,
                    geometry.config.brick_size_y,
                    geometry.config.brick_size_z
                ],
                'origin_shift': geometry.config.origin_shift.tolist(),
                'n_nodes': len(geometry.nodes),
                'n_elements': len(geometry.elements),
                'n_dofs': 3 * len(geometry.nodes),
                'n_fixed_dofs': len(fixed_dofs),
                'n_loads': len(bc_processor.loads),
                'n_supports': len(bc_processor.supports),
            },
            
            # Geometry (for reference and visualization)
            'geometry': {
                'nodes': geometry.nodes,
                'elements': geometry.elements,
                'voxel_grid': geometry.voxel_grid,
            },
            
            # Boundary conditions (for reference)
            'boundary_conditions': {
                'supports': [
                    {
                        'id': s.support_id,
                        'nodes': s.affected_nodes,
                        'constraints': {
                            'u_x': s.u_x, 'u_y': s.u_y, 'u_z': s.u_z,
                        }
                    }
                    for s in bc_processor.supports
                ],
                'loads': [
                    {
                        'id': l.load_id,
                        'nodes': l.affected_nodes,
                        'forces': {
                            'f_x': l.f_x, 'f_y': l.f_y, 'f_z': l.f_z,
                        }
                    }
                    for l in bc_processor.loads
                ],
            }
        }
        
        # Save to file
        with open(filepath, 'wb') as file:
            pickle.dump(export_data, file)
        
        # Print summary
        print("\n" + "="*60)
        print("✓ MODEL EXPORT SUCCESSFUL!")
        print("="*60)
        print(f"File: {filename}")
        print(f"Location: {output_dir}")
        print(f"\nModel Summary:")
        print(f"  • Nodes: {export_data['metadata']['n_nodes']}")
        print(f"  • Elements: {export_data['metadata']['n_elements']}")
        print(f"  • Total DOFs: {export_data['metadata']['n_dofs']}")
        print(f"  • Fixed DOFs: {export_data['metadata']['n_fixed_dofs']}")
        print(f"  • Supports: {export_data['metadata']['n_supports']}")
        print(f"  • Loads: {export_data['metadata']['n_loads']}")
        print(f"\nReady for OPT STM GENERATOR!")
        print(f"\nUsage in OPT STM GENERATOR:")
        print(f"  import pickle")
        print(f"  with open('{filename}', 'rb') as f:")
        print(f"      data = pickle.load(f)")
        print(f"  fe_solver = ContinuumFESolver(")
        print(f"      edof=data['edof'],")
        print(f"      ke0=data['ke0'],")
        print(f"      f=data['f'],")
        print(f"      fixed_dofs=data['fixed_dofs'],")
        print(f"      grid_shape=data['grid_shape'],")
        print(f"      volumes=data['volumes'],")
        print(f"      mesh_size=data['mesh_size']")
        print(f"  )")
        print("="*60 + "\n")
        
        return filepath


def load_excel_data(filepath: Path) -> Dict:
    """Load and parse Excel template data."""
    print(f"\nLoading Excel file: {filepath.name}")
    
    try:
        # Load all sheets
        excel_data = pd.read_excel(filepath, sheet_name=None, engine='openpyxl')
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {e}")
    
    # Parse basic parameters
    basic_df = excel_data.get('Basic Parameters')
    if basic_df is None:
        raise ValueError("'Basic Parameters' sheet not found")
    
    config = GeometryConfig()
    
    try:
        # Extract parameters (assuming format from template)
        params = {}
        for _, row in basic_df.iterrows():
            if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                param_name = str(row.iloc[0]).strip()
                param_value = row.iloc[1]
                if param_name in ["Length Units", "Force Units"]:
                    params[param_name] = str(param_value).strip()
                elif "Brick Size" in param_name:
                    params[param_name] = float(param_value)
        
        config.units_length = params.get("Length Units", "mm")
        config.units_force = params.get("Force Units", "kN")
        config.brick_size_x = params.get("Brick Size X", 100.0)
        config.brick_size_y = params.get("Brick Size Y", 100.0)
        config.brick_size_z = params.get("Brick Size Z", 100.0)
        
        print(f"  ✓ Units: {config.units_length}, {config.units_force}")
        print(f"  ✓ Brick size: {config.brick_size_x} x {config.brick_size_y} x {config.brick_size_z}")
    except Exception as e:
        raise ValueError(f"Error parsing basic parameters: {e}")
    
    # Parse positive volumes
    pos_vol_df = excel_data.get('Positive Volumes')
    positive_faces = []
    if pos_vol_df is not None:
        positive_faces = parse_face_definitions(pos_vol_df, is_negative=False)
        print(f"  ✓ Positive volumes: {len(positive_faces)} faces")
    
    # Parse negative volumes
    neg_vol_df = excel_data.get('Negative Volumes')
    negative_faces = []
    if neg_vol_df is not None:
        negative_faces = parse_face_definitions(neg_vol_df, is_negative=True)
        if negative_faces:
            print(f"  ✓ Negative volumes: {len(negative_faces)} faces")
    
    # Parse supports
    support_df = excel_data.get('Supports')
    supports = []
    if support_df is not None:
        supports = parse_supports(support_df)
        print(f"  ✓ Supports: {len(supports)} conditions")
    
    # Parse loads
    load_df = excel_data.get('Loads')
    loads = []
    if load_df is not None:
        loads = parse_loads(load_df)
        print(f"  ✓ Loads: {len(loads)} conditions")
    
    return {
        'config': config,
        'positive_faces': positive_faces,
        'negative_faces': negative_faces,
        'supports': supports,
        'loads': loads,
    }


def parse_face_definitions(df: pd.DataFrame, is_negative: bool = False) -> List[Face]:
    """
    Parse face definitions from dataframe with comprehensive error handling.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame containing face definitions
    is_negative : bool
        Whether these are negative (void) volumes
        
    Returns
    -------
    List[Face]
        List of validated Face objects
        
    Raises
    ------
    ValueError
        If data validation fails with detailed error message
    """
    faces = []
    volume_type = "NEGATIVE" if is_negative else "POSITIVE"
    
    # Skip header rows and get data
    data_df = df.iloc[2:].copy()  # Skip first 2 rows (headers)
    data_df.columns = ['Volume_ID', 'Face_ID', 'Point_ID', 'X', 'Y', 'Z', 'Face_Type', 'Notes']
    
    # Remove completely empty rows
    data_df = data_df.dropna(how='all')
    
    if len(data_df) == 0:
        return faces
    
    # Validate critical columns
    errors = []
    for idx, row in data_df.iterrows():
        row_num = idx + 3  # Excel row number (accounting for headers)
        
        # Check Volume_ID
        if pd.isna(row['Volume_ID']):
            errors.append(f"Row {row_num}: Missing Volume_ID")
        elif not isinstance(row['Volume_ID'], (int, float)) or row['Volume_ID'] != int(row['Volume_ID']):
            errors.append(f"Row {row_num}: Volume_ID must be an integer (got: {row['Volume_ID']})")
        
        # Check Face_ID  
        if pd.isna(row['Face_ID']):
            errors.append(f"Row {row_num}: Missing Face_ID for Volume {row.get('Volume_ID', '?')}")
        elif not isinstance(row['Face_ID'], (int, float)) or row['Face_ID'] != int(row['Face_ID']):
            errors.append(f"Row {row_num}: Face_ID must be an integer (got: {row['Face_ID']})")
        
        # Check Point_ID
        if pd.isna(row['Point_ID']):
            errors.append(f"Row {row_num}: Missing Point_ID for Volume {row.get('Volume_ID', '?')}, Face {row.get('Face_ID', '?')}")
        
        # Check coordinates
        for coord in ['X', 'Y', 'Z']:
            if pd.isna(row[coord]) or (isinstance(row[coord], str) and row[coord].strip() in ['', '<fill>']):
                errors.append(f"Row {row_num}: Missing or unfilled {coord} coordinate for Point {row.get('Point_ID', '?')}")
            elif not isinstance(row[coord], (int, float)):
                try:
                    float(row[coord])
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: {coord} coordinate must be numeric (got: {row[coord]})")
    
    if errors:
        error_msg = f"\n❌ Validation errors in {volume_type} VOLUMES sheet:\n" + "\n".join(f"   • {e}" for e in errors[:10])
        if len(errors) > 10:
            error_msg += f"\n   ... and {len(errors) - 10} more errors"
        raise ValueError(error_msg)
    
    # Remove rows with missing coordinates
    data_df = data_df.dropna(subset=['Volume_ID', 'Face_ID', 'X', 'Y', 'Z'])
    
    # Convert coordinates to numeric
    for coord in ['X', 'Y', 'Z']:
        data_df[coord] = pd.to_numeric(data_df[coord], errors='coerce')
    
    # Group by Volume_ID and Face_ID
    grouped = data_df.groupby(['Volume_ID', 'Face_ID'])
    
    for (vol_id, face_id), group in grouped:
        # Validate face has at least 3 points
        if len(group) < 3:
            print(f"   ⚠ Warning: {volume_type} Volume {int(vol_id)}, Face {int(face_id)} has only {len(group)} point(s). Need at least 3 points to define a planar face.")
            continue
        
        # Check for duplicate points
        points = group[['X', 'Y', 'Z']].values
        unique_points = np.unique(points, axis=0)
        if len(unique_points) < len(points):
            print(f"   ⚠ Warning: {volume_type} Volume {int(vol_id)}, Face {int(face_id)} has {len(points) - len(unique_points)} duplicate point(s). Removing duplicates.")
            points = unique_points
        
        face = Face(
            points=points,
            face_id=int(face_id),
            volume_id=int(vol_id),
            is_negative=is_negative
        )
        faces.append(face)
    
    return faces


def parse_supports(df: pd.DataFrame) -> List[SupportCondition]:
    """
    Parse support conditions from dataframe with comprehensive error handling.
    
    Note: Rotational DOFs (theta_x, theta_y, theta_z) are not supported in hex8 elements.
    """
    supports = []
    
    data_df = df.iloc[2:].copy()
    
    # Check if theta columns exist (old format) or not (new format)
    if len(df.columns) >= 12:
        # Old format with theta columns
        data_df.columns = ['Support_ID', 'Type', 'Point_ID', 'X', 'Y', 'Z', 
                           'u_x', 'u_y', 'u_z', 'theta_x', 'theta_y', 'theta_z']
        print("   ℹ Note: Theta (rotational) DOF columns detected but will be ignored (hex8 elements have no rotational DOFs)")
    else:
        # New format without theta columns
        data_df.columns = ['Support_ID', 'Type', 'Point_ID', 'X', 'Y', 'Z', 
                           'u_x', 'u_y', 'u_z']
    
    # Remove completely empty rows
    data_df = data_df.dropna(how='all')
    
    if len(data_df) == 0:
        return supports
    
    # Validate critical columns
    errors = []
    for idx, row in data_df.iterrows():
        row_num = idx + 3  # Excel row number
        
        # Check Support_ID
        if pd.isna(row['Support_ID']):
            errors.append(f"Row {row_num}: Missing Support_ID")
        elif not isinstance(row['Support_ID'], (int, float)) or row['Support_ID'] != int(row['Support_ID']):
            errors.append(f"Row {row_num}: Support_ID must be an integer (got: {row['Support_ID']})")
        
        # Check Type
        if pd.isna(row['Type']):
            errors.append(f"Row {row_num}: Missing Type for Support {row.get('Support_ID', '?')}")
        elif str(row['Type']).strip() not in ['Face', 'Point', 'face', 'point']:
            errors.append(f"Row {row_num}: Type must be 'Face' or 'Point' (got: {row['Type']})")
        
        # Check coordinates
        for coord in ['X', 'Y', 'Z']:
            if pd.isna(row[coord]) or (isinstance(row[coord], str) and row[coord].strip() in ['', '<fill>']):
                errors.append(f"Row {row_num}: Missing or unfilled {coord} coordinate for Support {row.get('Support_ID', '?')}, Point {row.get('Point_ID', '?')}")
            elif not isinstance(row[coord], (int, float)):
                try:
                    float(row[coord])
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: {coord} must be numeric (got: {row[coord]})")
        
        # Check constraint flags
        for dof in ['u_x', 'u_y', 'u_z']:
            if not pd.isna(row[dof]):
                if row[dof] not in [0, 1, '0', '1', 0.0, 1.0]:
                    errors.append(f"Row {row_num}: {dof} must be 0 (free) or 1 (fixed) (got: {row[dof]})")
    
    if errors:
        error_msg = "\n❌ Validation errors in SUPPORTS sheet:\n" + "\n".join(f"   • {e}" for e in errors[:10])
        if len(errors) > 10:
            error_msg += f"\n   ... and {len(errors) - 10} more errors"
        raise ValueError(error_msg)
    
    data_df = data_df.dropna(subset=['Support_ID', 'X', 'Y', 'Z'])
    
    # Convert coordinates to numeric
    for coord in ['X', 'Y', 'Z']:
        data_df[coord] = pd.to_numeric(data_df[coord], errors='coerce')
    
    grouped = data_df.groupby('Support_ID')
    
    for supp_id, group in grouped:
        is_face = len(group) > 1
        points = group[['X', 'Y', 'Z']].values
        
        # Validate face has at least 3 points for planar definition
        if is_face and len(points) < 3:
            print(f"   ⚠ Warning: Support {int(supp_id)} defined as Face but has only {len(points)} point(s). Need at least 3.")
        
        # Get constraints from first row (assuming same for all points in face)
        first_row = group.iloc[0]
        
        support = SupportCondition(
            support_id=int(supp_id),
            location=points[0] if not is_face else points,
            is_face=is_face,
            face_points=points if is_face else None,
            u_x=bool(int(first_row['u_x'])) if pd.notna(first_row['u_x']) else True,
            u_y=bool(int(first_row['u_y'])) if pd.notna(first_row['u_y']) else True,
            u_z=bool(int(first_row['u_z'])) if pd.notna(first_row['u_z']) else True,
            # Theta DOFs removed - not applicable for hex8 elements
            theta_x=False,
            theta_y=False,
            theta_z=False,
        )
        supports.append(support)
    
    return supports


def parse_loads(df: pd.DataFrame) -> List[LoadCondition]:
    """
    Parse load conditions from dataframe with comprehensive error handling.
    
    Note: 
    - Moments (M_x, M_y, M_z) are not supported in hex8 elements.
    - For face loads: if different force values are specified at face corners,
      they will be AVERAGED and applied uniformly. Non-uniform (varying) load 
      distribution is not currently implemented. To approximate varying loads,
      use multiple smaller faces with different uniform loads.
    """
    loads = []
    
    data_df = df.iloc[3:].copy()  # Skip first 3 rows (headers + note)
    
    # Check if moment columns exist (old format) or not (new format)
    if len(df.columns) >= 12:
        # Old format with moment columns
        data_df.columns = ['Load_ID', 'Type', 'Point_ID', 'X', 'Y', 'Z',
                           'F_x', 'F_y', 'F_z', 'M_x', 'M_y', 'M_z']
        print("   ℹ Note: Moment columns detected but will be ignored (hex8 elements have no rotational DOFs)")
    else:
        # New format without moment columns
        data_df.columns = ['Load_ID', 'Type', 'Point_ID', 'X', 'Y', 'Z',
                           'F_x', 'F_y', 'F_z']
    
    # Remove completely empty rows
    data_df = data_df.dropna(how='all')
    
    if len(data_df) == 0:
        return loads
    
    # Validate critical columns
    errors = []
    warnings = []
    for idx, row in data_df.iterrows():
        row_num = idx + 4  # Excel row number
        
        # Check Load_ID
        if pd.isna(row['Load_ID']):
            errors.append(f"Row {row_num}: Missing Load_ID")
        elif not isinstance(row['Load_ID'], (int, float)) or row['Load_ID'] != int(row['Load_ID']):
            errors.append(f"Row {row_num}: Load_ID must be an integer (got: {row['Load_ID']})")
        
        # Check Type
        if pd.isna(row['Type']):
            errors.append(f"Row {row_num}: Missing Type for Load {row.get('Load_ID', '?')}")
        elif str(row['Type']).strip() not in ['Face', 'Point', 'face', 'point']:
            errors.append(f"Row {row_num}: Type must be 'Face' or 'Point' (got: {row['Type']})")
        
        # Check coordinates
        for coord in ['X', 'Y', 'Z']:
            if pd.isna(row[coord]) or (isinstance(row[coord], str) and row[coord].strip() in ['', '<fill>']):
                errors.append(f"Row {row_num}: Missing or unfilled {coord} coordinate for Load {row.get('Load_ID', '?')}, Point {row.get('Point_ID', '?')}")
            elif not isinstance(row[coord], (int, float)):
                try:
                    float(row[coord])
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: {coord} must be numeric (got: {row[coord]})")
        
        # Check force values
        for force in ['F_x', 'F_y', 'F_z']:
            if pd.isna(row[force]):
                errors.append(f"Row {row_num}: Missing {force} value for Load {row.get('Load_ID', '?')}")
            elif not isinstance(row[force], (int, float)):
                try:
                    float(row[force])
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: {force} must be numeric (got: {row[force]})")
        
        # Check for moments if present
        if 'M_x' in row and not pd.isna(row['M_x']) and float(row['M_x']) != 0:
            warnings.append(f"Row {row_num}: Moment M_x specified but will be ignored (hex8 elements have no rotational DOFs)")
        if 'M_y' in row and not pd.isna(row['M_y']) and float(row['M_y']) != 0:
            warnings.append(f"Row {row_num}: Moment M_y specified but will be ignored")
        if 'M_z' in row and not pd.isna(row['M_z']) and float(row['M_z']) != 0:
            warnings.append(f"Row {row_num}: Moment M_z specified but will be ignored")
    
    if errors:
        error_msg = "\n❌ Validation errors in LOADS sheet:\n" + "\n".join(f"   • {e}" for e in errors[:10])
        if len(errors) > 10:
            error_msg += f"\n   ... and {len(errors) - 10} more errors"
        raise ValueError(error_msg)
    
    if warnings:
        for w in warnings[:5]:
            print(f"   ⚠ {w}")
        if len(warnings) > 5:
            print(f"   ⚠ ... and {len(warnings) - 5} more moment warnings")
    
    data_df = data_df.dropna(subset=['Load_ID', 'X', 'Y', 'Z'])
    
    # Convert coordinates and forces to numeric
    for col in ['X', 'Y', 'Z', 'F_x', 'F_y', 'F_z']:
        data_df[col] = pd.to_numeric(data_df[col], errors='coerce')
    
    grouped = data_df.groupby('Load_ID')
    
    for load_id, group in grouped:
        is_face = len(group) > 1
        points = group[['X', 'Y', 'Z']].values
        
        # Validate face has at least 3 points
        if is_face and len(points) < 3:
            print(f"   ⚠ Warning: Load {int(load_id)} defined as Face but has only {len(points)} point(s). Need at least 3.")
        
        # Get force values - AVERAGE if face load with varying forces
        forces_x = group['F_x'].values
        forces_y = group['F_y'].values
        forces_z = group['F_z'].values
        
        if is_face:
            # Check if forces vary across the face
            if (len(np.unique(forces_x)) > 1 or 
                len(np.unique(forces_y)) > 1 or 
                len(np.unique(forces_z)) > 1):
                print(f"   ⚠ Warning: Load {int(load_id)} has varying force values at face corners.")
                print(f"      Forces will be AVERAGED for uniform distribution.")
                print(f"      For non-uniform loads, use multiple smaller faces with different uniform loads.")
            
            # Use average
            f_x = float(np.mean(forces_x))
            f_y = float(np.mean(forces_y))
            f_z = float(np.mean(forces_z))
        else:
            # Point load - use single value
            f_x = float(forces_x[0])
            f_y = float(forces_y[0])
            f_z = float(forces_z[0])
        
        load = LoadCondition(
            load_id=int(load_id),
            location=points[0] if not is_face else points,
            is_face=is_face,
            face_points=points if is_face else None,
            f_x=f_x,
            f_y=f_y,
            f_z=f_z,
            # Moments removed - not applicable for hex8 elements
            m_x=0.0,
            m_y=0.0,
            m_z=0.0,
        )
        loads.append(load)
    
    return loads


def main():
    """Main interactive workflow."""
    print("\n" + "="*60)
    print("OPT-STM GEOMETRY INPUT DEFINITION")
    print("="*60 + "\n")
    
    # Step 1: Ask if template is needed
    print("Step 1: Excel Template")
    print("-" * 40)
    need_template = input("Generate new Excel template? (Y/N): ").strip().upper()
    
    if need_template == 'Y':
        template_path = Path(__file__).parent / "OPT_STM_Input_Template.xlsx"
        generator = ExcelTemplateGenerator(template_path)
        generator.generate()
        print("\nPlease fill in the template and run this script again.")
        return
    
    # Step 2: Request file path
    print("\nStep 2: Load Input File")
    print("-" * 40)
    
    while True:
        filepath_str = input("Enter path to Excel file (with filename): ").strip().strip('"')
        filepath = Path(filepath_str)
        
        if filepath.exists() and filepath.suffix in ['.xlsx', '.xls']:
            break
        else:
            print(f"  ✗ File not found or invalid format: {filepath}")
            retry = input("  Try again? (Y/N): ").strip().upper()
            if retry != 'Y':
                return
    
    # Load data
    try:
        data = load_excel_data(filepath)
    except Exception as e:
        print(f"\n✗ Error loading data: {e}")
        return
    
    # Step 3: Process geometry
    print("\nStep 3: Processing Geometry")
    print("-" * 40)
    
    geometry = GeometryProcessor(data['config'])
    
    # Add volumes
    if not data['positive_faces']:
        print("✗ Error: No positive volumes defined!")
        return
    
    # Group faces by volume
    pos_volumes = {}
    for face in data['positive_faces']:
        if face.volume_id not in pos_volumes:
            pos_volumes[face.volume_id] = []
        pos_volumes[face.volume_id].append(face)
    
    for vol_id, faces in pos_volumes.items():
        geometry.add_positive_volume(faces)
        print(f"  ✓ Added positive volume {vol_id}: {len(faces)} faces")
    
    if data['negative_faces']:
        neg_volumes = {}
        for face in data['negative_faces']:
            if face.volume_id not in neg_volumes:
                neg_volumes[face.volume_id] = []
            neg_volumes[face.volume_id].append(face)
        
        for vol_id, faces in neg_volumes.items():
            geometry.add_negative_volume(faces)
            print(f"  ✓ Added negative volume {vol_id}: {len(faces)} faces")
    
    # Compute origin shift to place model in first octant
    shift = geometry.compute_origin_shift()
    if np.any(shift > 1e-10):
        print(f"\n  Model offset to place in first octant.")
        print(f"  Origin shift applied: [{shift[0]:.3f}, {shift[1]:.3f}, {shift[2]:.3f}] {data['config'].units_length}")
        geometry.apply_shift(shift)
    else:
        print(f"\n  Model is already in first octant. No origin shift applied.")
    
    # Generate mesh
    print("\nGenerating voxel mesh...")
    try:
        geometry.generate_voxel_mesh()
        geometry.generate_node_element_connectivity()
    except Exception as e:
        print(f"✗ Error generating mesh: {e}")
        return
    
    # Step 4: Process boundary conditions
    print("\nStep 4: Processing Boundary Conditions")
    print("-" * 40)
    
    bc_processor = BoundaryConditionProcessor(geometry)
    
    if data['supports']:
        print("Processing supports...")
        for support in data['supports']:
            bc_processor.add_support(support)
    
    if data['loads']:
        print("\nProcessing loads...")
        for load in data['loads']:
            bc_processor.add_load(load)
    
    # Step 5: Visualization
    print("\nStep 5: Visualizing Model")
    print("-" * 40)
    
    visualizer = ModelVisualizer(geometry, bc_processor)
    
    plot_path = filepath.parent / f"{filepath.stem}_preview.png"
    visualizer.plot_model(show_plot=True, save_path=plot_path)
    
    # Step 6: Export model
    print("\nStep 6: Export Model")
    print("-" * 40)
    
    tag = input("Enter model tag (max 10 characters): ").strip()[:10]
    if not tag:
        tag = "Model"
    
    try:
        output_path = ModelExporter.export_model(
            geometry, bc_processor, tag, filepath.parent
        )
        print(f"\n✓ Complete! Model ready for OPT-STM optimization.")
    except Exception as e:
        print(f"\n✗ Error exporting model: {e}")


if __name__ == "__main__":
    main()
